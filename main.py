from fastapi import FastAPI, File, UploadFile, Form, Depends, HTTPException, status, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pathlib import Path
from pydantic import BaseModel, EmailStr
from typing import List, Optional, Dict, Any, Tuple
import os
import io
import re
import json
import logging
import secrets
import base64
import smtplib
import time
import concurrent.futures
from datetime import datetime, timedelta, timezone
from urllib.parse import quote_plus, urlparse
from zipfile import ZipFile, ZIP_DEFLATED
from functools import lru_cache
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv

import joblib
import pandas as pd
import requests
from groq import Groq
from PyPDF2 import PdfReader
from sqlalchemy.orm.session import Session
from sqlalchemy import text, or_
from openpyxl import Workbook

from services.feature_analyzer import build_complete_feature_vector, FEATURE_COLUMNS as DEFAULT_FEATURE_COLUMNS
from services.database import engine, get_db, Base
from services.models import User, AnalysisResult, QuizResult, Job, TpoLogin, InterestedJob, ShortlistedJob, Notification, JobResult
from services.quiz_generator import generate_quiz_questions
from services.auth import (
    hash_password,
    verify_password,
    create_access_token,
    get_current_user,
    get_current_student,
    get_current_tpo,
)

# Logger Setup
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(filename="backend_debug.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)
BASE_DIR = Path(__file__).resolve().parent

# Roadmap skill artifacts (.pkl) may live one level above this folder.
# Prefer current directory, fall back to parent if needed.
MODEL_DIR = BASE_DIR
if not (MODEL_DIR / "roadmap_model.pkl").exists():
    parent = BASE_DIR.parent
    if (parent / "roadmap_model.pkl").exists():
        MODEL_DIR = parent

# Ensure EMAIL_* / SMTP_* vars from .env are available in this module.
load_dotenv(dotenv_path=BASE_DIR / ".env")

# ── Create database tables on startup ─────────────────────────────────────────
Base.metadata.create_all(bind=engine)


def ensure_db_schema_compatibility() -> None:
    """Backfill columns required by current ORM models on older databases."""
    with engine.begin() as conn:
        # TPO login table columns introduced after initial schema.
        conn.execute(
            text(
                'ALTER TABLE "TPO_login" ADD COLUMN IF NOT EXISTS password_reset_token VARCHAR(100)'
            )
        )
        conn.execute(
            text(
                'ALTER TABLE "TPO_login" ADD COLUMN IF NOT EXISTS password_reset_expires TIMESTAMP WITH TIME ZONE'
            )
        )
        conn.execute(
            text("ALTER TABLE \"TPO_login\" ADD COLUMN IF NOT EXISTS first_name VARCHAR(100) DEFAULT ''")
        )
        conn.execute(
            text("ALTER TABLE \"TPO_login\" ADD COLUMN IF NOT EXISTS last_name VARCHAR(100) DEFAULT ''")
        )

        # Users table profile and password-reset columns added over time.
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS moodle_id VARCHAR(8)'))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS first_name VARCHAR(100) DEFAULT ''"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS last_name VARCHAR(100) DEFAULT ''"))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS year VARCHAR(4)'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS division VARCHAR(1)'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS semester INTEGER'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS sgpa DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS marks_10th DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS marks_12th DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS diploma_avg DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS sem1 DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS sem2 DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS sem3 DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS sem4 DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS sem5 DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS sem6 DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS atkt_count INTEGER DEFAULT 0'))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS atkt_subjects TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS drop_year VARCHAR(3) DEFAULT 'No'"))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS internships JSON'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS projects JSON'))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS core_interests TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS core_skills TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS github_profile TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS linkedin_profile TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS achievements TEXT DEFAULT ''"))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS resume_score DOUBLE PRECISION'))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS resume_filename VARCHAR(255) DEFAULT ''"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS resume_url TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS resume_text TEXT DEFAULT ''"))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS password_reset_token VARCHAR(100)'))
        conn.execute(text('ALTER TABLE users ADD COLUMN IF NOT EXISTS password_reset_expires TIMESTAMP WITH TIME ZONE'))
        # Profile photo URL storage
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS photo_url TEXT DEFAULT ''"))

        # Jobs table columns introduced for richer posting details.
        conn.execute(text("ALTER TABLE jobs ADD COLUMN IF NOT EXISTS job_role VARCHAR(255) DEFAULT ''"))
        conn.execute(text('ALTER TABLE jobs ADD COLUMN IF NOT EXISTS min_cgpa DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE jobs ADD COLUMN IF NOT EXISTS min_resume_score DOUBLE PRECISION'))
        conn.execute(text("ALTER TABLE jobs ADD COLUMN IF NOT EXISTS required_certifications TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE jobs ADD COLUMN IF NOT EXISTS preferred_skills TEXT DEFAULT ''"))
        conn.execute(text('ALTER TABLE jobs ADD COLUMN IF NOT EXISTS package_lpa DOUBLE PRECISION'))
        conn.execute(text('ALTER TABLE jobs ADD COLUMN IF NOT EXISTS salary DOUBLE PRECISION'))
        conn.execute(text("ALTER TABLE jobs ADD COLUMN IF NOT EXISTS deadline VARCHAR(100) DEFAULT ''"))
        conn.execute(text("ALTER TABLE jobs ADD COLUMN IF NOT EXISTS company_logo TEXT DEFAULT ''"))
        conn.execute(text('UPDATE jobs SET package_lpa = salary WHERE package_lpa IS NULL AND salary IS NOT NULL'))
        conn.execute(text('UPDATE jobs SET salary = package_lpa WHERE salary IS NULL AND package_lpa IS NOT NULL'))

        # Shortlisted mapping table metadata.
        conn.execute(text("ALTER TABLE shortlisted_jobs ADD COLUMN IF NOT EXISTS source VARCHAR(20) DEFAULT 'manual'"))
        conn.execute(text("UPDATE shortlisted_jobs SET source = 'manual' WHERE source IS NULL"))

        # Job results table fields for auto-broadcasted round updates.
        conn.execute(text("ALTER TABLE job_results ADD COLUMN IF NOT EXISTS result_status VARCHAR(20) DEFAULT 'Qualified'"))
        conn.execute(text("ALTER TABLE job_results ADD COLUMN IF NOT EXISTS remarks TEXT DEFAULT ''"))


ensure_db_schema_compatibility()

# ── Global model placeholders ───────────────────────────────────────────────
feature_columns = DEFAULT_FEATURE_COLUMNS
roadmap_model = None
roadmap_mlb = None
roadmap_vectorizer = None
roadmap_models_loaded = False

def load_ml_models_background():
    global roadmap_model, roadmap_mlb, roadmap_vectorizer, roadmap_models_loaded
    logger.info("Loading ML models in background thread...")
    try:
        start_time = time.time()

        roadmap_model = joblib.load(MODEL_DIR / "roadmap_model.pkl")
        roadmap_mlb = joblib.load(MODEL_DIR / "roadmap_mlb.pkl")
        roadmap_vectorizer = joblib.load(MODEL_DIR / "roadmap_vectorizer.pkl")
        
        roadmap_models_loaded = True
        logger.info(f"ML models loaded successfully in {time.time() - start_time:.2f} seconds.")
    except Exception as e:
        logger.error(f"Critical error loading ML models: {e}")

import threading
threading.Thread(target=load_ml_models_background, daemon=True).start()



# ----------------- Helper: URL and YouTube validation -----------------
@lru_cache(maxsize=1000)
def is_valid_url(url: str) -> bool:
    try:
        if not url or not isinstance(url, str):
            logger.debug("is_valid_url: empty or non-string URL")
            return False
        parsed = urlparse(url)
        if parsed.scheme not in ("http", "https"):
            logger.debug(f"is_valid_url: invalid scheme for {url}")
            return False
        
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"}
        
        # Try HEAD first as it's faster
        try:
            r = requests.head(url, allow_redirects=True, timeout=3, headers=headers)
            if 200 <= r.status_code < 400:
                logger.info(f"is_valid_url: VALID (HEAD {r.status_code}) - {url}")
                return True
            else:
                logger.info(f"is_valid_url: INVALID (HEAD {r.status_code}) - {url}")
        except Exception as e:
            logger.debug(f"is_valid_url: HEAD exception for {url}: {e}")
            
        # Fallback to GET for sites that block HEAD
        try:
            r = requests.get(url, allow_redirects=True, timeout=3, stream=True, headers=headers)
            if 200 <= r.status_code < 400:
                logger.info(f"is_valid_url: VALID (GET {r.status_code}) - {url}")
                return True
            logger.info(f"is_valid_url: INVALID (GET {r.status_code}) - {url}")
        except Exception as e:
            logger.debug(f"is_valid_url: GET exception for {url}: {e}")
            
        return False
    except Exception as e:
        logger.error(f"is_valid_url: unexpected error for {url}: {e}")
        return False


@lru_cache(maxsize=1000)
def is_valid_youtube_id(video_id: str) -> bool:
    try:
        if not video_id or not isinstance(video_id, str):
            return False
        if len(video_id) != 11:
            return False
        oembed = f"https://www.youtube.com/oembed?url=http://www.youtube.com/watch?v={video_id}&format=json"
        r = requests.get(oembed, timeout=3)
        return r.status_code == 200
    except Exception:
        return False


def google_search_fallback(query: str) -> str:
    return f"https://www.google.com/search?q={quote_plus(query)}"


def sanitize_roadmap_resources(roadmap_data: dict, topic_name: str) -> dict:
    """Shared helper to validate and fill roadmap resources with direct links.
    Ensures that we don't return 0 counts and that links are DIRECT wherever possible.
    """
    if not isinstance(roadmap_data, dict):
        return roadmap_data

    sanitized = {
        "nodes": roadmap_data.get("nodes", []),
        "edges": roadmap_data.get("edges", []),
        "courses": [],
        "certificates": [],
        "youtube": []
    }
    TARGETS = {"courses": 3, "certificates": 2, "youtube": 3}

    def process_course(c: Any) -> Optional[Dict[str, Any]]:
        if not isinstance(c, dict): return None
        url = c.get("url")
        if url and is_valid_url(url): return c
        # Try to resolve direct link if missing or invalid
        title = c.get("title", "")
        platform = c.get("platform", "")
        if not title: return None
        query = f"{topic_name} {title} {platform} course".strip()
        resolved = resolve_direct_link(query)
        if resolved:
            c_copy = dict(c)
            c_copy["url"] = resolved
            return c_copy
        return None

    def process_cert(cert: Any) -> Optional[Dict[str, Any]]:
        if not isinstance(cert, dict): return None
        url = cert.get("url")
        if url and is_valid_url(url): return cert
        title = cert.get("title", "")
        provider = cert.get("provider", "")
        if not title: return None
        query = f"{topic_name} {title} {provider} certification".strip()
        resolved = resolve_direct_link(query)
        if resolved:
            cert_copy = dict(cert)
            cert_copy["url"] = resolved
            return cert_copy
        return None

    def process_youtube(v: Any) -> Optional[Dict[str, Any]]:
        if not isinstance(v, dict): return None
        vid = v.get("videoId")
        if vid and is_valid_youtube_id(vid): return v
        title = v.get("title", "")
        if not title: return None
        query = f"{topic_name} {title} tutorial".strip()
        resolved_vid = resolve_youtube_video_id(query)
        if resolved_vid:
            v_copy = dict(v)
            v_copy["videoId"] = resolved_vid
            return v_copy
        return None

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        fut_c = [executor.submit(process_course, c) for c in (roadmap_data.get("courses") or [])]
        fut_cert = [executor.submit(process_cert, c) for c in (roadmap_data.get("certificates") or [])]
        fut_yt = [executor.submit(process_youtube, v) for v in (roadmap_data.get("youtube") or [])]

        for f in concurrent.futures.as_completed(fut_c):
            if res := f.result(): sanitized["courses"].append(res)
        for f in concurrent.futures.as_completed(fut_cert):
            if res := f.result(): sanitized["certificates"].append(res)
        for f in concurrent.futures.as_completed(fut_yt):
            if res := f.result(): sanitized["youtube"].append(res)

    # Fallback logic: if still below targets, generate generic items and resolve them
    def add_fallbacks(key, count, query_suffix, item_factory):
        while len(sanitized[key]) < count:
            idx = len(sanitized[key]) + 1
            title = f"{topic_name} {key[:-1].capitalize()} {idx}"
            query = f"{topic_name} {title} {query_suffix}".strip()
            
            # 1. Try resolving a direct link
            resolved = resolve_direct_link(query) if key != "youtube" else resolve_youtube_video_id(query)
            
            # 2. If resolution fails, construct a platform-specific search link
            # This redirects the user directly to the target website's internal search.
            if not resolved:
                q = quote_plus(query)
                if key == "courses":
                    resolved = f"https://www.udemy.com/courses/search/?q={q}"
                elif key == "certificates":
                    resolved = f"https://www.coursera.org/search?query={q}"

            if resolved:
                item = item_factory(title, resolved)
                sanitized[key].append(item)
            else:
                # Last resort for YouTube if no ID found
                if key == "youtube":
                    sanitized[key].append({
                        "title": title, 
                        "channel": "YouTube", 
                        "videoId": None, 
                        "description": "Click to search for this tutorial."
                    })
                else:
                    break

    add_fallbacks("courses", TARGETS["courses"], "online course", 
                  lambda t, r: {"title": t, "platform": "Udemy", "url": r, "level": "Beginner", "description": "Recommended learning resource."})
    add_fallbacks("certificates", TARGETS["certificates"], "certification",
                  lambda t, r: {"title": t, "provider": "Coursera", "url": r, "description": "Professional certification path."})
    add_fallbacks("youtube", TARGETS["youtube"], "tutorial", 
                  lambda t, r: {"title": r if isinstance(r, str) and len(r) == 11 else t, "channel": "YouTube", "videoId": r if isinstance(r, str) and len(r) == 11 else None, "description": "Helpful overview video."})

    return sanitized


def _http_get_text(url: str, timeout: int = 8) -> Optional[str]:
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; HireReadyBot/1.0)"}
        r = requests.get(url, headers=headers, timeout=timeout)
        if r.status_code == 200:
            return r.text
    except Exception:
        pass
    return None


def resolve_direct_link(query: str) -> Optional[str]:
    """Try to find a direct link for the given query.

    Strategy:
    1. If Bing API key is configured, use Bing Web Search (recommended).
    2. Else if Google Custom Search key + CX configured, use Google Custom Search.
    3. Fallback to DuckDuckGo HTML scraping.

    Returns the first HTTPS URL found, or None.
    """
    if not query:
        return None

    # 1) Bing Web Search (preferred if key present)
    bing_key = os.getenv("BING_SEARCH_KEY")
    bing_endpoint = os.getenv("BING_SEARCH_ENDPOINT", "https://api.bing.microsoft.com/v7.0/search")
    if bing_key:
        try:
            headers = {"Ocp-Apim-Subscription-Key": bing_key}
            r = requests.get(bing_endpoint, params={"q": query, "count": 1}, headers=headers, timeout=8)
            if r.status_code == 200:
                j = r.json()
                url = j.get("webPages", {}).get("value", [{}])[0].get("url")
                if url and is_valid_url(url):
                    return url
        except Exception:
            pass

    # 2) Google Custom Search
    g_key = os.getenv("GOOGLE_SEARCH_KEY")
    g_cx = os.getenv("GOOGLE_CX")
    if g_key and g_cx:
        try:
            r = requests.get(
                "https://www.googleapis.com/customsearch/v1",
                params={"key": g_key, "cx": g_cx, "q": query, "num": 1},
                timeout=8,
            )
            if r.status_code == 200:
                j = r.json()
                items = j.get("items") or []
                if items:
                    link = items[0].get("link")
                    if link and is_valid_url(link):
                        return link
        except Exception:
            pass

    # 3) DuckDuckGo HTML fallback
    dd_url = f"https://duckduckgo.com/html/?q={quote_plus(query)}"
    html = _http_get_text(dd_url)
    if not html:
        return None

    # Look for common anchor pattern for results: <a class="result__a" href="https://...">
    m = re.search(r"<a[^>]+class=\"[^\"]*result__a[^\"]*\"[^>]+href=\"([^\"]+)\"", html)
    if m:
        href = m.group(1)
        # Some DDG links can be redirects (uddg=...). If so, try to extract the uddg parameter.
        uddg_m = re.search(r"uddg=([^&]+)", href)
        if uddg_m:
            try:
                decoded = requests.utils.unquote(uddg_m.group(1))
                if is_valid_url(decoded):
                    return decoded
            except Exception:
                pass
        if is_valid_url(href):
            return href

    # Fallback: extract first https link in page
    m2 = re.search(r'https?://[^" ]+', html)
    if m2:
        candidate = m2.group(0)
        if is_valid_url(candidate):
            return candidate
    return None


def resolve_youtube_video_id(query: str) -> Optional[str]:
    """Return a YouTube videoId for the query using the YouTube Data API when available,
    otherwise fall back to scraping the YouTube search page.
    """
    if not query:
        return None

    yt_key = os.getenv("YOUTUBE_API_KEY")
    if yt_key:
        try:
            r = requests.get(
                "https://www.googleapis.com/youtube/v3/search",
                params={"part": "snippet", "q": query, "type": "video", "maxResults": 1, "key": yt_key},
                timeout=8,
            )
            if r.status_code == 200:
                j = r.json()
                items = j.get("items") or []
                if items:
                    vid = items[0].get("id", {}).get("videoId")
                    if vid and is_valid_youtube_id(vid):
                        return vid
        except Exception:
            pass

    # Fallback: scrape YouTube search results page
    yt_search = f"https://www.youtube.com/results?search_query={quote_plus(query)}"
    html = _http_get_text(yt_search)
    if not html:
        return None

    # look for /watch?v=VIDEOID
    m = re.search(r"/watch\?v=([A-Za-z0-9_-]{11})", html)
    if m:
        vid = m.group(1)
        if is_valid_youtube_id(vid):
            return vid
    return None


# ----------------- Simple in-memory TTL cache for generated resources -----------------
RESOURCE_CACHE_TTL = int(os.getenv("RESOURCE_CACHE_TTL_SECONDS", "86400"))  # seconds; default 24h
RESOURCE_CACHE: Dict[str, Tuple[float, Any]] = {}  # maps cache_key -> (timestamp, value)


def _cache_get(key: str):
    entry = RESOURCE_CACHE.get(key)
    if not entry:
        return None
    ts, val = entry
    if (time.time() - ts) > RESOURCE_CACHE_TTL:
        RESOURCE_CACHE.pop(key, None)
        return None
    return val


def _cache_set(key: str, value):
    RESOURCE_CACHE[key] = (time.time(), value)


# Precompute/Cache Skills for Roadmap
ROLE_SKILLS_CACHE = {}
try:
    df_roles = pd.read_csv(BASE_DIR / "technical_roles_and_skills.csv")
    for _, row in df_roles.iterrows():
        # Map normalized role name to list of skills
        role_name = str(row["Role"]).strip().lower()
        skills_list = [s.strip() for s in str(row["Key Skills"]).split(";") if s.strip()]
        ROLE_SKILLS_CACHE[role_name] = skills_list
    logger.info(f"Loaded {len(ROLE_SKILLS_CACHE)} roles into skill cache.")
except Exception as e:
    logger.error(f"Failed to load technical_roles_and_skills.csv: {e}")

app = FastAPI()


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    """Catch-all so every error returns JSON, never plain text."""
    logger.exception("Unhandled error on %s %s", request.method, request.url.path)
    return JSONResponse(
        status_code=500,
        content={"detail": "Internal server error. Please try again later."},
    )


# Paths for serving the built frontend
FRONTEND_DIST = BASE_DIR / "frontend" / "dist"
FRONTEND_INDEX = FRONTEND_DIST / "index.html"
FRONTEND_ASSETS = FRONTEND_DIST / "assets"
UPLOADS_DIR = BASE_DIR / "uploads"
RESUMES_DIR = UPLOADS_DIR / "resumes"
RESUMES_DIR.mkdir(parents=True, exist_ok=True)
RESULTS_DIR = UPLOADS_DIR / "results"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

# Allow CORS for the React dev server
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount built asset folder if present
if FRONTEND_ASSETS.exists():
    app.mount("/assets", StaticFiles(directory=FRONTEND_ASSETS), name="assets")
else:
    # Optional fallback: serve public assets if dist not built yet
    PUBLIC_DIR = BASE_DIR / "frontend" / "public"
    if PUBLIC_DIR.exists():
        app.mount("/assets", StaticFiles(directory=PUBLIC_DIR), name="assets")

if UPLOADS_DIR.exists():
    app.mount("/uploads", StaticFiles(directory=UPLOADS_DIR), name="uploads")


@app.get("/", response_class=HTMLResponse)
async def serve_root(request: Request):
    """Serve the built HireReady frontend."""
    index_path = FRONTEND_INDEX if FRONTEND_INDEX.exists() else BASE_DIR / "frontend" / "index.html"
    if not index_path.exists():
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Frontend build not found. Run 'npm install' then 'npm run build' inside frontend/.",
        )
    return FileResponse(index_path)


@app.get("/reset-password", response_class=HTMLResponse)
async def serve_reset_password_page(request: Request):
    """Serve frontend app for reset-password route (SPA entry)."""
    index_path = FRONTEND_INDEX if FRONTEND_INDEX.exists() else BASE_DIR / "frontend" / "index.html"
    if not index_path.exists():
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Frontend build not found. Run 'npm install' then 'npm run build' inside frontend/.",
        )
    return FileResponse(index_path)
    
@app.get("/vite.svg")
async def serve_vite_svg():
    svg_path = FRONTEND_DIST / "vite.svg"
    if not svg_path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="vite.svg not found. Build frontend.")
    return FileResponse(svg_path)


@app.get("/api/ping")
def ping():
    """Simple health endpoint for the UI button."""
    return {"status": "ok"}


# ═══════════════════════════════════════════════════════════════════════════════
# AUTH ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

class RegisterRequest(BaseModel):
    name: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: EmailStr
    password: str
    role: str = "student"  # "student" or "tpo"


class LoginRequest(BaseModel):
    email: str
    password: str


class ForgotPasswordRequest(BaseModel):
    email: str


class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str


class NotifyShortlistedRequest(BaseModel):
    student_ids: List[str] = []
    job_id: Optional[str] = None


class DownloadSelectedResumesRequest(BaseModel):
    student_ids: List[str] = []
    job_id: Optional[str] = None


class ShortlistStudentRequest(BaseModel):
    job_id: str
    student_id: str


class ShortlistAllRequest(BaseModel):
    job_id: str
    student_ids: List[str] = []


class ProfilePhotoRequest(BaseModel):
    # The storage file path (recommended) or a public URL
    filePath: str


@app.post('/api/profile/photo')
def save_profile_photo(payload: ProfilePhotoRequest, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    """Persist a user's profile photo path/url after upload to storage.

    The frontend should upload directly to Supabase (or another bucket) and then POST
    the saved `filePath` (or public URL) here. The server verifies the object exists
    when possible and stores the path in the users table under `photo_url`.
    """
    file_path = (payload.filePath or '').strip()
    if not file_path:
        raise HTTPException(status_code=400, detail='filePath is required')

    # If SUPABASE_URL is configured and file_path does not look like a full URL,
    # check the public object URL for existence. For private buckets, backend
    # verification with the service role key is recommended (not required here).
    supabase_url = os.getenv('SUPABASE_URL')
    public_url = file_path
    try:
        if supabase_url and not re.match(r'^https?://', file_path):
            public_url = f"{supabase_url}/storage/v1/object/public/profile_photos/{file_path}"

        # HEAD is faster; fallback to GET if HEAD not allowed
        r = requests.head(public_url, allow_redirects=True, timeout=6)
        if r.status_code >= 400:
            r2 = requests.get(public_url, allow_redirects=True, timeout=6, stream=True)
            if r2.status_code >= 400:
                raise HTTPException(status_code=400, detail='Uploaded file not found in storage')
    except requests.RequestException:
        # If supabase isn't configured or request failed, reject to be safe.
        if supabase_url:
            raise HTTPException(status_code=400, detail='Failed to verify uploaded file')

    # Update the user record
    user = db.query(User).filter(User.id == current_user.id).first()
    if not user:
        raise HTTPException(status_code=404, detail='User not found')

    # Attempt to remove previous photo from Supabase storage (if present and service key available).
    try:
        prev = (user.photo_url or '').strip()
        supabase_service_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
        supabase_url_cfg = os.getenv('SUPABASE_URL')
        # Determine previous file path within profile_photos bucket
        prev_path = None
        if prev:
            if prev.startswith('http') and '/profile_photos/' in prev:
                prev_path = prev.split('/profile_photos/', 1)[1]
            elif prev.startswith('profile_photos/'):
                prev_path = prev[len('profile_photos/'):]
            else:
                # If stored earlier as just a path, accept it
                prev_path = prev

        # Only attempt deletion when service key + supabase URL available and prev_path exists
        if prev_path and supabase_service_key and supabase_url_cfg:
            # Avoid deleting the same file we are about to save
            if prev_path != file_path:
                delete_url = requests.utils.requote_uri(f"{supabase_url_cfg}/storage/v1/object/profile_photos/{prev_path}")
                headers_del = {"apikey": supabase_service_key, "Authorization": f"Bearer {supabase_service_key}"}
                try:
                    dresp = requests.delete(delete_url, headers=headers_del, timeout=8)
                    if dresp.status_code not in (200, 204):
                        logger.info("Failed to delete previous profile photo (%s): %s", delete_url, dresp.status_code)
                except Exception:
                    logger.exception("Error while deleting previous profile photo")
    except Exception:
        logger.exception("Unexpected error when attempting to delete previous profile photo")

    # Store the public URL for ease of use in the frontend; store the path if you prefer.
    user.photo_url = public_url
    db.add(user)
    db.commit()

    return {"ok": True, "photo_url": user.photo_url}

RESULT_STATUSES = {"Selected", "Rejected", "Qualified"}


@app.post("/api/auth/register")
def register(data: RegisterRequest, db: Session = Depends(get_db)):
    """Create a new account. Students → users table, TPOs → TPO_login table."""
    email = data.email.strip().lower()
    role = data.role.strip().lower()
    if role not in ("student", "tpo"):
        raise HTTPException(status_code=400, detail="Role must be 'student' or 'tpo'.")

    first_name = (data.first_name or "").strip()
    last_name = (data.last_name or "").strip()
    full_name = (data.name or f"{first_name} {last_name}").strip()
    if not full_name:
        raise HTTPException(status_code=400, detail="Name is required.")

    # ── TPO Registration → TPO_login table ────────────────────────────────
    if role == "tpo":
        existing_tpo = db.query(TpoLogin).filter(TpoLogin.email == email).first()
        if existing_tpo:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A TPO account with this email already exists.",
            )

        tpo = TpoLogin(
            email=email,
            first_name=first_name,
            last_name=last_name,
            password=hash_password(data.password),
        )
        db.add(tpo)
        db.commit()

        import uuid as _uuid
        tpo_uuid = str(_uuid.uuid5(_uuid.NAMESPACE_DNS, email))
        token = create_access_token(email, role="tpo")

        return {
            "token": token,
            "user": {
                "id": tpo_uuid,
                "name": full_name,
                "first_name": first_name,
                "last_name": last_name,
                "email": tpo.email,
                "role": "tpo",
            },
        }

    # ── Student Registration → users table ────────────────────────────────
    existing = db.query(User).filter(User.email == email).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="An account with this email already exists.",
        )

    user = User(
        name=full_name,
        first_name=first_name,
        last_name=last_name,
        email=email,
        password_hash=hash_password(data.password),
        role="student",
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    token = create_access_token(str(user.id), role="student")

    return {
        "token": token,
        "user": {
            "id": str(user.id),
            "name": user.name,
            "first_name": user.first_name or "",
            "last_name": user.last_name or "",
            "email": user.email,
            "role": user.role,
            "github_username": user.github_username or "",
            "leetcode_username": user.leetcode_username or "",
            "resume_filename": user.resume_filename or "",
            "mobile_number": user.mobile_number or "",
            "cgpa": user.cgpa,
            "certifications": user.certifications or "",
            "preferred_job_roles": user.preferred_job_roles or "",
        },
    }


@app.post("/api/auth/login")
def login(data: LoginRequest, db: Session = Depends(get_db)):
    """Authenticate. Checks TPO_login first, then users table."""
    email = data.email.strip().lower()

    # ── Try TPO_login table first ─────────────────────────────────────────
    tpo = db.query(TpoLogin).filter(TpoLogin.email == email).first()
    if tpo:
        if not verify_password(data.password, tpo.password):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid email or password.",
            )
        import uuid as _uuid
        tpo_uuid = str(_uuid.uuid5(_uuid.NAMESPACE_DNS, tpo.email))
        token = create_access_token(tpo.email, role="tpo")
        return {
            "token": token,
            "user": {
                "id": tpo_uuid,
                "name": ((f"{(tpo.first_name or '').strip()} {(tpo.last_name or '').strip()}").strip() or tpo.email),
                "first_name": tpo.first_name or "",
                "last_name": tpo.last_name or "",
                "email": tpo.email,
                "role": "tpo",
            },
        }

    # ── Fall back to users (student) table ────────────────────────────────
    user = db.query(User).filter(User.email == email).first()

    if not user or not verify_password(data.password, user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid email or password.",
        )

    token = create_access_token(str(user.id), role=user.role)

    return {
        "token": token,
        "user": {
            "id": str(user.id),
            "name": user.name,
            "first_name": user.first_name or "",
            "last_name": user.last_name or "",
            "email": user.email,
            "role": user.role,
            "github_username": user.github_username or "",
            "leetcode_username": user.leetcode_username or "",
            "resume_filename": user.resume_filename or "",
            "mobile_number": user.mobile_number or "",
            "cgpa": user.cgpa,
            "certifications": user.certifications or "",
            "preferred_job_roles": user.preferred_job_roles or "",
        },
    }


# ═══════════════════════════════════════════════════════════════════════════════
# FORGOT / RESET PASSWORD
# ═══════════════════════════════════════════════════════════════════════════════

def validate_password_strength(password: str) -> list:
    """Return list of unmet password rules."""
    errors = []
    if len(password) < 8:
        errors.append("Minimum 8 characters")
    if not re.search(r'[A-Z]', password):
        errors.append("At least one uppercase letter")
    if not re.search(r'[a-z]', password):
        errors.append("At least one lowercase letter")
    if not re.search(r'[0-9]', password):
        errors.append("At least one number")
    if not re.search(r'[!@#$%^&*()_+\-=\[\]{};\':"\\|,.<>\/?]', password):
        errors.append("At least one special character")
    return errors


def resolve_frontend_base_url(request: Optional[Request] = None) -> str:
    """Return the best frontend base URL for reset links across environments."""
    configured = (os.getenv("FRONTEND_BASE_URL", "") or "").strip().rstrip("/")
    if configured:
        return configured

    if request is not None:
        origin = (request.headers.get("origin") or "").strip().rstrip("/")
        if origin.startswith("http://") or origin.startswith("https://"):
            return origin

        referer = (request.headers.get("referer") or "").strip()
        if referer:
            parsed = urlparse(referer)
            if parsed.scheme in {"http", "https"} and parsed.netloc:
                return f"{parsed.scheme}://{parsed.netloc}".rstrip("/")

        forwarded_host = (request.headers.get("x-forwarded-host") or "").strip()
        if forwarded_host:
            forwarded_proto = (request.headers.get("x-forwarded-proto") or request.url.scheme or "http").strip()
            return f"{forwarded_proto}://{forwarded_host}".rstrip("/")

        return str(request.base_url).rstrip("/")

    return "http://localhost:5173"


def _first_env(*names: str, default: str = "") -> str:
    """Return first non-empty environment variable from a list of names."""
    for name in names:
        value = (os.getenv(name, "") or "").strip()
        if value:
            return value
    return default


def _env_bool(*names: str, default: bool = False) -> bool:
    """Parse a boolean from env vars using common true/false strings."""
    value = _first_env(*names)
    if not value:
        return default
    return value.lower() in {"1", "true", "yes", "on"}


def send_password_reset_email(recipient_email: str, token: str, request: Optional[Request] = None) -> bool:
    """Send a password reset link to the given email using SMTP."""
    smtp_host = _first_env("SMTP_HOST", "MAIL_HOST", "MAIL_SERVER", default="smtp.gmail.com")
    smtp_port = int(_first_env("SMTP_PORT", "MAIL_PORT", default="587"))
    email_user = _first_env(
        "EMAIL_USER",
        "EMAIL_USERNAME",
        "SMTP_USERNAME",
        "MAIL_USERNAME",
        "MAIL_USER",
    )
    email_pass = _first_env(
        "EMAIL_PASS",
        "EMAIL_PASSWORD",
        "SMTP_PASSWORD",
        "MAIL_PASSWORD",
        "APP_PASSWORD",
    )
    email_from = _first_env("EMAIL_FROM", "MAIL_FROM", default="HireReady")
    frontend_base_url = resolve_frontend_base_url(request)
    smtp_use_ssl = _env_bool("SMTP_USE_SSL", "MAIL_USE_SSL", default=(smtp_port == 465))
    smtp_use_tls = _env_bool("SMTP_USE_TLS", "MAIL_USE_TLS", default=(not smtp_use_ssl))

    if not email_user or not email_pass:
        # SMTP not configured: don't treat this as an error for the
        # forgot-password endpoint. Log the reset link so developers
        # can copy it during local development / testing.
        reset_link = f"{frontend_base_url}/reset-password?token={token}"
        logger.warning(
            "SMTP credentials missing; set EMAIL_USER/EMAIL_PASS (or SMTP_USERNAME/SMTP_PASSWORD, MAIL_USERNAME/MAIL_PASSWORD). Logging reset link instead of sending email."
        )
        logger.info("Password reset link for %s: %s", recipient_email, reset_link)
        return True

    # Format the 'From' header for branding (e.g., "HireReady <user@gmail.com>")
    if "@" in email_from:
        display_from = email_from
    else:
        display_from = f"{email_from} <{email_user}>"

    reset_link = f"{frontend_base_url}/reset-password?token={token}"
    body = (
        "Hello,\n\n"
        "We received a request to reset your HireReady password.\n\n"
        f"Click the link below to set a new password:\n{reset_link}\n\n"
        "This link expires in 15 minutes.\n\n"
        "If you did not request this, you can safely ignore this email.\n\n"
        "- HireReady Team"
    )

    msg = MIMEMultipart()
    msg["Subject"] = "Password Reset - HireReady"
    msg["From"] = display_from
    msg["To"] = recipient_email
    msg.attach(MIMEText(body, "plain"))

    try:
        smtp_client = smtplib.SMTP_SSL if smtp_use_ssl else smtplib.SMTP
        with smtp_client(smtp_host, smtp_port, timeout=20) as server:
            server.ehlo()
            if smtp_use_tls and not smtp_use_ssl:
                server.starttls()
                server.ehlo()
            server.login(email_user, email_pass)
            # CRITICAL: Envelope sender MUST be a valid email address (email_user)
            server.sendmail(email_user, [recipient_email], msg.as_string())
        logger.info("Password reset email sent to %s", recipient_email)
        return True
    except Exception as e:
        logger.error("Failed to send password reset email: %s", str(e))
        # Do NOT raise here. Email delivery failures should not crash the
        # forgot-password endpoint — we return a generic success message
        # to avoid revealing account existence and to keep the endpoint
        # resilient to SMTP issues.
        return False


@app.post("/api/auth/forgot-password")
def forgot_password(data: ForgotPasswordRequest, request: Request, db: Session = Depends(get_db)):
    """Generate reset token and email link, without revealing account existence."""
    email = data.email.strip().lower()
    generic_message = "If the email exists, a password reset link has been sent."
    token = secrets.token_urlsafe(32)
    expiry = datetime.now(timezone.utc) + timedelta(minutes=15)
    try:
        user = db.query(User).filter(User.email == email).first()
        if user:
            user.password_reset_token = token
            user.password_reset_expires = expiry
            db.commit()
            try:
                send_password_reset_email(email, token, request=request)
            except Exception:
                # send_password_reset_email now returns False on failure,
                # but keep defensive logging in case of unexpected errors.
                logger.exception("Failed to send password reset email (student) to %s", email)
            return {"message": generic_message}

        tpo = db.query(TpoLogin).filter(TpoLogin.email == email).first()
        if tpo:
            tpo.password_reset_token = token
            tpo.password_reset_expires = expiry
            db.commit()
            try:
                send_password_reset_email(email, token, request=request)
            except Exception:
                logger.exception("Failed to send password reset email (tpo) to %s", email)
            return {"message": generic_message}

        return {"message": generic_message}
    except Exception as exc:
        # Catch anything unexpected (DB disconnected, schema error, etc.)
        logger.exception("Error processing forgot-password for %s: %s", email, exc)
        # Still return the generic message so the UI doesn't receive a 500.
        return {"message": generic_message}


@app.post("/api/auth/reset-password")
def reset_password(data: ResetPasswordRequest, db: Session = Depends(get_db)):
    """Reset password using a valid token."""
    # Validate password strength
    errors = validate_password_strength(data.new_password)
    if errors:
        raise HTTPException(
            status_code=400,
            detail="Password must contain: " + ", ".join(errors),
        )

    now = datetime.now(timezone.utc)

    # Check student table
    user = db.query(User).filter(
        User.password_reset_token == data.token,
    ).first()
    if user:
        if user.password_reset_expires and user.password_reset_expires > now:
            user.password_hash = hash_password(data.new_password)
            user.password_reset_token = None
            user.password_reset_expires = None
            db.commit()
            return {"message": "Password reset successfully. You can now log in."}
        else:
            raise HTTPException(status_code=400, detail="Reset token has expired. Please request a new one.")

    # Check TPO table
    tpo = db.query(TpoLogin).filter(
        TpoLogin.password_reset_token == data.token,
    ).first()
    if tpo:
        if tpo.password_reset_expires and tpo.password_reset_expires > now:
            tpo.password = hash_password(data.new_password)
            tpo.password_reset_token = None
            tpo.password_reset_expires = None
            db.commit()
            return {"message": "Password reset successfully. You can now log in."}
        else:
            raise HTTPException(status_code=400, detail="Reset token has expired. Please request a new one.")

    raise HTTPException(status_code=400, detail="Invalid reset token.")


@app.get("/api/auth/me")
def get_me(current_user: User = Depends(get_current_user)):
    """Return the current user's profile."""
    return {
        "id": str(current_user.id),
        "name": current_user.name,
        "email": current_user.email,
        "role": current_user.role,
        "github_username": current_user.github_username or "",
        "leetcode_username": current_user.leetcode_username or "",
        "resume_filename": current_user.resume_filename or "",
        "mobile_number": current_user.mobile_number or "",
        "cgpa": current_user.cgpa,
        "certifications": current_user.certifications or "",
        "preferred_job_roles": current_user.preferred_job_roles or "",
        "moodle_id": current_user.moodle_id or "",
        "prn_no": current_user.prn_no or "",
        "year": current_user.year or "",
        "division": current_user.division or "",
        "semester": current_user.semester,
        "sgpa": current_user.sgpa,
        "marks_10th": current_user.marks_10th,
        "marks_12th": current_user.marks_12th,
        "diploma_avg": current_user.diploma_avg,
        "sem1": current_user.sem1,
        "sem2": current_user.sem2,
        "sem3": current_user.sem3,
        "sem4": current_user.sem4,
        "sem5": current_user.sem5,
        "sem6": current_user.sem6,
        "atkt_count": current_user.atkt_count,
        "atkt_subjects": current_user.atkt_subjects or "",
        "drop_year": current_user.drop_year or "No",
        "internships": current_user.internships or [],
        "projects": current_user.projects or [],
        "core_interests": current_user.core_interests or "",
        "core_skills": current_user.core_skills or "",
        "github_profile": current_user.github_profile or "",
        "linkedin_profile": current_user.linkedin_profile or "",
        "achievements": current_user.achievements or "",
        "resume_url": current_user.resume_url or "",
        "photo_url": current_user.photo_url or "",
    }


SKILL_FEATURE_KEYS = set(DEFAULT_FEATURE_COLUMNS[:45])
PROJECT_FEATURE_KEYS = [
    "num_backend_projects",
    "num_ai_projects",
    "num_mobile_projects",
    "num_cloud_projects",
    "num_security_projects",
]


def _extract_first_json_object(text: str) -> Optional[Dict[str, Any]]:
    """Extract the first JSON object from an LLM response string."""
    if not text:
        return None

    try:
        return json.loads(text)
    except Exception:
        pass

    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or start >= end:
        return None

    try:
        return json.loads(text[start:end + 1])
    except Exception:
        return None


def build_structured_candidate_data(
    feature_vector: Dict[str, Any],
    resume_text: str,
    **_kwargs,
) -> Dict[str, Any]:
    """Convert extracted features into the structured payload consumed by the LLM.

    Extra keyword arguments (e.g. legacy github_username, leetcode_username)
    are silently ignored for backward compatibility.
    """
    skills = sorted(
        [
            key
            for key, value in feature_vector.items()
            if key in SKILL_FEATURE_KEYS and isinstance(value, (int, float)) and value > 0
        ]
    )

    project_counts = {
        "backend": int(feature_vector.get("num_backend_projects", 0) or 0),
        "ai": int(feature_vector.get("num_ai_projects", 0) or 0),
        "mobile": int(feature_vector.get("num_mobile_projects", 0) or 0),
        "cloud": int(feature_vector.get("num_cloud_projects", 0) or 0),
        "security": int(feature_vector.get("num_security_projects", 0) or 0),
    }
    projects = [f"{k} projects ({v})" for k, v in project_counts.items() if v > 0]

    # Internship domains detected from resume
    internship_domains = [
        key.replace("internship_", "").title()
        for key in ["internship_backend", "internship_ai", "internship_cloud",
                    "internship_security", "internship_mobile", "internship_data"]
        if feature_vector.get(key, 0) > 0
    ]

    return {
        "skills": skills,
        "projects": projects,
        "internship_domains": internship_domains,
        "project_counts": project_counts,
        "resume_excerpt": (resume_text or "")[:1200],
    }


def _fallback_prediction(extracted_data: Dict[str, Any], feature_vector: Dict[str, Any]) -> Dict[str, Any]:
    """Deterministic fallback if the LLM is unavailable.

    Scoring: skills 50%, projects 30%, internships 20%.
    """
    skills_count = len(extracted_data.get("skills", []))
    projects_total = sum(int(v) for v in extracted_data.get("project_counts", {}).values())
    internship_count = len(extracted_data.get("internship_domains", []))

    score = 0.0
    score += min(skills_count / 20.0, 1.0) * 50.0  # skills 50%
    score += min(projects_total / 5.0, 1.0) * 30.0  # projects 30%
    score += min(internship_count / 3.0, 1.0) * 20.0  # internships 20%
    score = round(max(0.0, min(100.0, score)), 2)

    suggested_role = "Software Engineer"
    if extracted_data.get("project_counts", {}).get("ai", 0) > 0:
        suggested_role = "ML Engineer"
    elif extracted_data.get("project_counts", {}).get("mobile", 0) > 0:
        suggested_role = "Mobile Developer"
    elif extracted_data.get("project_counts", {}).get("cloud", 0) > 0:
        suggested_role = "Cloud Engineer"
    elif extracted_data.get("project_counts", {}).get("backend", 0) > 0:
        suggested_role = "Backend Developer"

    recommended_roles = [
        {"role": suggested_role, "score": score},
        {"role": "Software Engineer", "score": max(0.0, round(score - 6, 2))},
        {"role": "Full Stack Developer", "score": max(0.0, round(score - 10, 2))},
    ]

    weaknesses = []
    if skills_count < 8:
        weaknesses.append("Limited technical skill breadth")
    if projects_total < 3:
        weaknesses.append("Insufficient project volume")
    if internship_count == 0:
        weaknesses.append("No relevant internship experience detected")
    if not weaknesses:
        weaknesses.append("No major weaknesses detected from available signals")

    suggestions = []
    if skills_count < 8:
        suggestions.append("Learn more in-demand technologies to broaden your skill set")
    if projects_total < 3:
        suggestions.append("Build 2 to 3 production-style portfolio projects")
    if internship_count == 0:
        suggestions.append("Pursue a relevant internship to gain industry experience")
    if not suggestions:
        suggestions.append("Keep improving system design and interview communication")

    return {
        "score": score,
        "role": recommended_roles[0]["role"] if recommended_roles else "Software Engineer",
        "recommended_roles": recommended_roles,
        "weakness": weaknesses,
        "suggestions": suggestions,
        "source": "fallback",
    }


def get_ai_prediction(extracted_data: Dict[str, Any], feature_vector: Dict[str, Any]) -> Dict[str, Any]:
    """Predict readiness and role using Groq with structured candidate data."""
    api_key = (os.getenv("GROQ_API_KEY") or "").strip()
    if not api_key:
        logger.warning("GROQ_API_KEY missing; using fallback prediction.")
        return _fallback_prediction(extracted_data, feature_vector)

    model_name = (os.getenv("GROQ_READINESS_MODEL") or "llama-3.3-70b-versatile").strip()
    client = Groq(api_key=api_key)

    prompt = (
        "You are an expert placement evaluator. "
        "Evaluate the candidate based solely on their resume skills, projects, and internship experience. "
        "Return ONLY valid JSON with this exact schema: "
        "{\"score\": number, \"role\": string, \"recommended_roles\": [{\"role\": string, \"score\": number}], "
        "\"weakness\": [string], \"suggestions\": [string]}. "
        "Rules: score must be between 0 and 100, recommended_roles max 3 entries, weakness/suggestions max 5 each.\n\n"
        f"Candidate Data:\n{json.dumps(extracted_data, ensure_ascii=True)}"
    )

    try:
        completion = client.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": "Return strict JSON only."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            max_tokens=500,
        )

        raw = (completion.choices[0].message.content or "").strip()
        parsed = _extract_first_json_object(raw)
        if not parsed:
            logger.warning("AI prediction parse failed; using fallback prediction.")
            return _fallback_prediction(extracted_data, feature_vector)

        score = float(parsed.get("score", 0))
        score = round(max(0.0, min(100.0, score)), 2)

        role = str(parsed.get("role") or "").strip()
        recommended_roles = parsed.get("recommended_roles") or []
        normalized_roles = []
        for item in recommended_roles[:3]:
            if not isinstance(item, dict):
                continue
            r = str(item.get("role") or "").strip()
            if not r:
                continue
            try:
                s = round(float(item.get("score", 0)), 2)
            except Exception:
                s = 0.0
            normalized_roles.append({"role": r, "score": max(0.0, min(100.0, s))})

        if not normalized_roles:
            normalized_roles = _fallback_prediction(extracted_data, feature_vector)["recommended_roles"]

        if not role:
            role = normalized_roles[0]["role"] if normalized_roles else "Software Engineer"

        weakness = parsed.get("weakness") if isinstance(parsed.get("weakness"), list) else []
        suggestions = parsed.get("suggestions") if isinstance(parsed.get("suggestions"), list) else []
        weakness = [str(x).strip() for x in weakness if str(x).strip()][:5]
        suggestions = [str(x).strip() for x in suggestions if str(x).strip()][:5]

        if not weakness:
            weakness = ["Unable to infer weaknesses confidently from provided data"]
        if not suggestions:
            suggestions = ["Improve weak areas with consistent weekly practice"]

        return {
            "score": score,
            "role": role,
            "recommended_roles": normalized_roles,
            "weakness": weakness,
            "suggestions": suggestions,
            "source": "groq",
        }
    except Exception as exc:
        logger.error("Groq readiness prediction failed: %s", exc)
        return _fallback_prediction(extracted_data, feature_vector)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: Analysis Pipeline
# ═══════════════════════════════════════════════════════════════════════════════

def run_analysis_pipeline(
    user: User,
    db: Session,
    resume_text: str = "",
    **_kwargs,
):
    """
    Core logic to extract features, run model, and save result.
    Call this whenever profile details change.

    Extra keyword arguments (e.g. legacy github_username, leetcode_username)
    are silently ignored for backward compatibility.
    """
    # If no resume text provided, try to use saved
    if not resume_text:
        resume_text = user.resume_text or ""

    # 1. Extract feature vector (resume-only)
    feature_vector = build_complete_feature_vector(
        resume_text=resume_text,
    )

    # 2. AI readiness + role prediction using structured extracted data.
    extracted_data = build_structured_candidate_data(
        feature_vector=feature_vector,
        resume_text=resume_text,
    )
    ai_prediction = get_ai_prediction(extracted_data, feature_vector)
    readiness = float(ai_prediction.get("score", 0))

    # 4. Categorise
    if readiness >= 75:
        category = "Placement Ready"
    elif readiness >= 50:
        category = "Almost Ready"
    else:
        category = "Needs Improvement"

    # 5. Role recommendation from AI
    role_list = ai_prediction.get("recommended_roles", [])

    # 6. Calculate category-specific scores (scale 0-10)
    def calculate_category_scores(fvec, u):
        missing_info = {
            "edu": [], "skill": [], "contact": [], "intern": [], "exp": [], "proj": []
        }
        
        # Education: Scale CGPA to 10
        edu = min(10.0, (u.cgpa or 0.0))
        if not u.cgpa:
            missing_info["edu"].append("CGPA not found in profile")
        elif u.cgpa < 8.0:
            missing_info["edu"].append("CGPA is below 8.0")
        
        # Skills: 20+ detected skills = 10/10
        skill_keys = [
            "Python", "Java", "C++", "JavaScript", "SQL", "Node", "React", "NextJS",
            "Docker", "Kubernetes", "CI/CD", "AWS", "NLP", "LLM", "SystemDesign", "DBMS"
        ]
        
        # Calculate score based on core skills
        core_detected_skills = [k for k in skill_keys if fvec.get(k, 0) > 0]
        skill_score = min(10.0, len(core_detected_skills) / 1.5) # Roughly 15 skills for 10/10
        
        # Display ALL extracted skills, not just core scoring skills
        all_detected_skills = [k for k, v in fvec.items() if k in SKILL_FEATURE_KEYS and isinstance(v, (int, float)) and v > 0]
        
        missing_skills = [k for k in skill_keys if k not in core_detected_skills]
        if len(core_detected_skills) < 10:
            missing_info["skill"].extend(missing_skills[:3])
        
        # Contact: Name, Email (always present), Mobile, LinkedIn
        contact = 5.0 # baseline for registered user
        if u.mobile_number: 
            contact += 2.5
        else:
            missing_info["contact"].append("Mobile Number")
            
        if u.linkedin_profile: 
            contact += 2.5
        else:
            missing_info["contact"].append("LinkedIn Profile")
        
        # Internships: 1=5, 2=8, 3+=10
        intern_keys = ["internship_backend", "internship_ai", "internship_cloud", "internship_security", "internship_mobile", "internship_data"]
        intern_count = sum(1 for k in intern_keys if fvec.get(k, 0) > 0)
        intern_score = 10.0 if intern_count >= 3 else (8.0 if intern_count == 2 else (5.0 if intern_count == 1 else 0.0))
        
        if intern_count == 0:
            missing_info["intern"].append("Relevant Internship")
        elif intern_count < 2:
            missing_info["intern"].append("Second Internship")
        
        # Projects: 5+ projects = 10/10
        proj_keys = ["num_backend_projects", "num_ai_projects", "num_mobile_projects", "num_cloud_projects", "num_security_projects"]
        proj_count = sum(fvec.get(k, 0) for k in proj_keys)
        # Cap count from individual domains to avoid single-domain bloating
        effective_proj_count = sum(min(fvec.get(k, 0), 2) for k in proj_keys)
        proj_score = min(10.0, effective_proj_count * 2.5)
        
        if proj_count < 3:
            missing_info["proj"].append(f"At least {3 - proj_count} more projects")
        
        # Experience: Weighted average of projects and internships for now
        exp_score = (intern_score * 0.6) + (proj_score * 0.4)
        if exp_score < 7:
            missing_info["exp"].append("More hands-on industry experience")
        
        return {
            "edu": round(edu, 1),
            "skill": round(skill_score, 1),
            "contact": round(contact, 1),
            "intern": round(intern_score, 1),
            "exp": round(exp_score, 1),
            "proj": round(proj_score, 1),
            "skills_list": sorted(all_detected_skills),
            "missing": missing_info
        }

    cat_scores = calculate_category_scores(feature_vector, user)

    # 7. Generate AI Suggestions
    from services.suggestion_engine import generate_suggestions
    engine_suggestions = generate_suggestions(
        readiness_score=readiness,
        features=feature_vector,
        categories_scores=cat_scores,
        resume_text=resume_text
    )
    ai_suggestions = {
        "llm": {
            "weakness": ai_prediction.get("weakness", []),
            "suggestions": ai_prediction.get("suggestions", []),
            "role": ai_prediction.get("role", ""),
            "source": ai_prediction.get("source", "fallback"),
        },
        "rules": engine_suggestions,
    }

    # 8. Save result
    try:
        analysis = AnalysisResult(
            user_id=user.id,
            resume_text_preview=resume_text[:200] if resume_text else "",
            github_username="",
            leetcode_username="",
            features=json.dumps(feature_vector) if isinstance(feature_vector, dict) else feature_vector,
            readiness_score=readiness,
            readiness_category=category,
            recommended_roles=role_list,
            # Re-assigned scores
            education_score=cat_scores["edu"],
            skills_score=cat_scores["skill"],
            contact_score=cat_scores["contact"],
            internship_score=cat_scores["intern"],
            experience_score=cat_scores["exp"],
            project_score=cat_scores["proj"],
            ai_suggestions=ai_suggestions,
            missing_details=cat_scores.get("missing", {})
        )
        db.add(analysis)

        # Store resume_score on user for shortlisting
        user.resume_score = readiness

        db.commit()
        db.refresh(analysis)
        logger.info("Auto-analysis saved: %s", analysis.id)
        return analysis
    except Exception as exc:
        logger.error("Failed to save analysis: %s", exc)
        db.rollback()
        return None

# ═══════════════════════════════════════════════════════════════════════════════
# PROFILE ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.put("/api/auth/profile")
async def update_profile(
    name: Optional[str] = Form(None),
    github_username: Optional[str] = Form(None),
    leetcode_username: Optional[str] = Form(None),
    mobile_number: Optional[str] = Form(None),
    cgpa: Optional[str] = Form(None),
    certifications: Optional[str] = Form(None),
    preferred_job_roles: Optional[str] = Form(None),
    moodle_id: Optional[str] = Form(None),
    prn_no: Optional[str] = Form(None),
    year: Optional[str] = Form(None),
    division: Optional[str] = Form(None),
    semester: Optional[int] = Form(None),
    sgpa: Optional[float] = Form(None),
    marks_10th: Optional[str] = Form(None),
    marks_12th: Optional[str] = Form(None),
    diploma_avg: Optional[str] = Form(None),
    sem1: Optional[str] = Form(None),
    sem2: Optional[str] = Form(None),
    sem3: Optional[str] = Form(None),
    sem4: Optional[str] = Form(None),
    sem5: Optional[str] = Form(None),
    sem6: Optional[str] = Form(None),
    atkt_count: Optional[int] = Form(None),
    atkt_subjects: Optional[str] = Form(None),
    drop_year: Optional[str] = Form(None),
    internships: Optional[str] = Form(None), # JSON string
    projects: Optional[str] = Form(None),    # JSON string
    core_interests: Optional[str] = Form(None),
    core_skills: Optional[str] = Form(None),
    github_profile: Optional[str] = Form(None),
    linkedin_profile: Optional[str] = Form(None),
    achievements: Optional[str] = Form(None),
    resume: UploadFile = File(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Update the current user's profile details."""
    new_resume_text = None

    if name is not None:
        current_user.name = name.strip()
    
    if github_username is not None:
        current_user.github_username = github_username.strip()

    if leetcode_username is not None:
        current_user.leetcode_username = leetcode_username.strip()

    if mobile_number is not None:
        current_user.mobile_number = mobile_number.strip()

    if cgpa is not None and cgpa.strip():
        try:
            current_user.cgpa = float(cgpa.strip())
        except ValueError:
            pass

    if certifications is not None:
        current_user.certifications = certifications.strip()

    if preferred_job_roles is not None:
        current_user.preferred_job_roles = preferred_job_roles.strip()

    if moodle_id is not None: current_user.moodle_id = moodle_id.strip()
    if prn_no is not None: current_user.prn_no = prn_no.strip()
    if year is not None: current_user.year = year.strip()
    if division is not None: current_user.division = division.strip()
    if semester is not None: current_user.semester = semester
    if sgpa is not None: current_user.sgpa = sgpa

    def _parse_optional_float(value: Optional[str]) -> Optional[float]:
        if value is None:
            return None
        value = value.strip()
        if value == "":
            return None
        try:
            return float(value)
        except ValueError:
            return None

    if marks_10th is not None: current_user.marks_10th = _parse_optional_float(marks_10th)
    if marks_12th is not None: current_user.marks_12th = _parse_optional_float(marks_12th)
    if diploma_avg is not None: current_user.diploma_avg = _parse_optional_float(diploma_avg)
    if sem1 is not None: current_user.sem1 = _parse_optional_float(sem1)
    if sem2 is not None: current_user.sem2 = _parse_optional_float(sem2)
    if sem3 is not None: current_user.sem3 = _parse_optional_float(sem3)
    if sem4 is not None: current_user.sem4 = _parse_optional_float(sem4)
    if sem5 is not None: current_user.sem5 = _parse_optional_float(sem5)
    if sem6 is not None: current_user.sem6 = _parse_optional_float(sem6)

    # Auto-calculate CGPA from available semester SGPIs.
    # This keeps CGPA consistent with semester-wise academic inputs.
    sgpi_values = [
        current_user.sem1,
        current_user.sem2,
        current_user.sem3,
        current_user.sem4,
        current_user.sem5,
        current_user.sem6,
    ]
    valid_sgpi_values = [v for v in sgpi_values if v is not None]
    if valid_sgpi_values:
        current_user.cgpa = round(sum(valid_sgpi_values) / len(valid_sgpi_values), 2)
    else:
        current_user.cgpa = None

    if atkt_count is not None: current_user.atkt_count = atkt_count
    if atkt_subjects is not None: current_user.atkt_subjects = atkt_subjects.strip()
    if drop_year is not None: current_user.drop_year = drop_year.strip()
    if core_interests is not None: current_user.core_interests = core_interests.strip()
    if core_skills is not None: current_user.core_skills = core_skills.strip()
    if github_profile is not None: current_user.github_profile = github_profile.strip()
    if linkedin_profile is not None: current_user.linkedin_profile = linkedin_profile.strip()
    if achievements is not None: current_user.achievements = achievements.strip()

    if internships is not None:
        try:
            current_user.internships = json.loads(internships)
        except:
            pass
    if projects is not None:
        try:
            current_user.projects = json.loads(projects)
        except:
            pass

    # Handle resume upload
    if resume and resume.filename:
        try:
            filename = resume.filename
            pdf_bytes = await resume.read() # Use await for async file read
            
            # Extract text using PyPDF2
            reader = PdfReader(io.BytesIO(pdf_bytes))
            text = ""
            for page in reader.pages:
                extracted = page.extract_text()
                if extracted:
                    text += extracted + "\n"

            safe_suffix = Path(filename).name.replace(" ", "_")
            stored_name = f"{current_user.id}_{int(datetime.now(timezone.utc).timestamp())}_{safe_suffix}"
            stored_path = RESUMES_DIR / stored_name
            stored_path.write_bytes(pdf_bytes)
            
            # Save strictly necessary fields
            current_user.resume_filename = filename
            current_user.resume_url = f"/uploads/resumes/{stored_name}"
            current_user.resume_text = text  # Store text for analysis
            
            new_resume_text = text
            logger.info("Updated resume for user %s: %s", current_user.email, filename)
        except Exception as e:
            logger.error("Failed to process resume upload: %s", e)
            raise HTTPException(status_code=400, detail="Invalid PDF file")

    db.commit()
    db.refresh(current_user)

    # 3. Trigger Auto-Analysis
    # We pass the NEW text if uploaded, otherwise None (helper uses saved)
    latest_analysis = run_analysis_pipeline(
        user=current_user,
        db=db,
        resume_text=new_resume_text if new_resume_text else None,
        # Helper will fetch saved usernames from User object
    )

    return {
        "user": {
            "id": str(current_user.id),
            "name": current_user.name,
            "resume_filename": current_user.resume_filename,
            "resume_url": current_user.resume_url or "",
            "photo_url": current_user.photo_url or "",
            "email": current_user.email,
            "github_username": current_user.github_username,
            "leetcode_username": current_user.leetcode_username,
            "mobile_number": current_user.mobile_number or "",
            "cgpa": current_user.cgpa,
            "certifications": current_user.certifications or "",
            "preferred_job_roles": current_user.preferred_job_roles or "",
            "moodle_id": current_user.moodle_id or "",
            "prn_no": current_user.prn_no or "",
            "year": current_user.year or "",
            "division": current_user.division or "",
            "semester": current_user.semester,
            "sgpa": current_user.sgpa,
            "marks_10th": current_user.marks_10th,
            "marks_12th": current_user.marks_12th,
            "diploma_avg": current_user.diploma_avg,
            "sem1": current_user.sem1,
            "sem2": current_user.sem2,
            "sem3": current_user.sem3,
            "sem4": current_user.sem4,
            "sem5": current_user.sem5,
            "sem6": current_user.sem6,
            "atkt_count": current_user.atkt_count,
            "atkt_subjects": current_user.atkt_subjects or "",
            "drop_year": current_user.drop_year or "No",
            "internships": current_user.internships or [],
            "projects": current_user.projects or [],
            "core_interests": current_user.core_interests or "",
            "core_skills": current_user.core_skills or "",
            "github_profile": current_user.github_profile or "",
            "linkedin_profile": current_user.linkedin_profile or "",
            "achievements": current_user.achievements or "",
        },
        "analysis": {
            "status": "success",
            "readiness_score": latest_analysis.readiness_score,
            "readiness_category": latest_analysis.readiness_category,
            "recommended_roles": latest_analysis.recommended_roles,
            "education_score": latest_analysis.education_score,
            "skills_score": latest_analysis.skills_score,
            "contact_score": latest_analysis.contact_score,
            "internship_score": latest_analysis.internship_score,
            "experience_score": latest_analysis.experience_score,
            "project_score": latest_analysis.project_score,
            "ai_suggestions": latest_analysis.ai_suggestions,
            "missing_details": latest_analysis.missing_details,
            "total_features_used": len([v for v in (json.loads(latest_analysis.features).values() if isinstance(latest_analysis.features, str) else latest_analysis.features.values() if latest_analysis.features else {}) if v > 0]),
            "skills_list": sorted([k for k, v in (json.loads(latest_analysis.features) if isinstance(latest_analysis.features, str) else latest_analysis.features or {}).items() if k in SKILL_FEATURE_KEYS and isinstance(v, (int, float)) and v > 0]),
            "created_at": latest_analysis.created_at
        } if latest_analysis else None
    }


@app.get("/api/analysis/latest")
def get_latest_analysis(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get the most recent analysis result for the user."""
    # TPOs don't have analysis results
    if current_user.role == "tpo":
        return {"status": "no_analysis"}

    # Query AnalysisResult table
    latest = (
        db.query(AnalysisResult)
        .filter(AnalysisResult.user_id == current_user.id)
        .order_by(AnalysisResult.created_at.desc())
        .first()
    )

    if not latest:
        return {"status": "no_analysis"}

    return {
        "status": "success",
        "readiness_score": latest.readiness_score,
        "readiness_category": latest.readiness_category,
        "recommended_roles": latest.recommended_roles,
        "education_score": latest.education_score,
        "skills_score": latest.skills_score,
        "contact_score": latest.contact_score,
        "internship_score": latest.internship_score,
        "experience_score": latest.experience_score,
        "project_score": latest.project_score,
        "ai_suggestions": latest.ai_suggestions,
        "missing_details": latest.missing_details,
        "total_features_used": len([v for v in (json.loads(latest.features).values() if isinstance(latest.features, str) else latest.features.values() if latest.features else {}) if v > 0]),
        "skills_list": sorted([k for k, v in (json.loads(latest.features) if isinstance(latest.features, str) else latest.features or {}).items() if k in SKILL_FEATURE_KEYS and isinstance(v, (int, float)) and v > 0]),
        "created_at": latest.created_at
    }


# NOTE: analyze-full-profile is now deprecated/redundant effectively,
# but can be kept as a direct tool if needed. 



# ═══════════════════════════════════════════════════════════════════════════════
# ANALYSIS ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

class StudentFeatures(BaseModel):
    features: dict


class FeatureExtractionRequest(BaseModel):
    resume_text: str = ""


@app.post("/api/extract-features")
def extract_features(data: FeatureExtractionRequest):
    """
    Extract a 56-feature dictionary from resume text.
    The output is directly usable as input to the /analyze-student endpoint.
    """
    feature_vector = build_complete_feature_vector(
        resume_text=data.resume_text,
    )
    return {"features": feature_vector}


class SkillPredictionRequest(BaseModel):
    role: str


@app.post("/api/predict-skills")
def predict_skills(data: SkillPredictionRequest):
    """
    Phase 1: Predict core tech skills for a job title.
    Returns an array of strings (Skill Bubbles).
    """
    role = data.role.strip()
    if not role:
        raise HTTPException(status_code=400, detail="Role name is required")
    
    # 1. Check Cache First (Instant)
    normalized_role = role.lower()
    if normalized_role in ROLE_SKILLS_CACHE:
        return {
            "role": role,
            "skills": sorted(ROLE_SKILLS_CACHE[normalized_role]),
            "source": "cache"
        }
    
    # 2. Fallback to ML Model (Predictive)
    if not roadmap_models_loaded or roadmap_model is None:
        logger.warning("ML Prediction skipped: models not yet loaded.")
        return {
            "role": role,
            "skills": ["Loading models...", "Please wait"],
            "note": "AI models are still initializing in the background. Please try again in 30 seconds."
        }

    try:
        v = roadmap_vectorizer.transform([role])
        p = roadmap_model.predict(v)
        skills = roadmap_mlb.inverse_transform(p)[0] # Extract first result
        
        return {
            "role": role,
            "skills": sorted(list(skills)),
            "source": "ml_model"
        }
    except Exception as e:
        logger.error(f"ML Prediction error: {e}")
        raise HTTPException(status_code=500, detail="Failed to predict skills")


class LearningPathRequest(BaseModel):
    skill: str


@app.post("/api/generate-learning-path")
def generate_learning_path(data: LearningPathRequest):
    """
    Phase 2: Generate a step-by-step learning path using Groq AI.
    Returns nodes and edges for React Flow.
    """
    skill = data.skill.strip()
    if not skill:
        raise HTTPException(status_code=400, detail="Skill name is required")

    cache_key = f"roadmap:skill:{skill.lower()}"
    cached = _cache_get(cache_key)
    if cached:
        logger.info(f"Serving cached roadmap for skill: {skill}")
        return cached

    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured")

    client = Groq(api_key=api_key)

    system_prompt = (
        "You are a professional curriculum designer. Generate a learning roadmap for the given skill. "
        "Return a STRICT JSON object with these keys:\n"
        "1. 'nodes': array of 6-8 objects with 'id' (like 'n1','n2') and 'label' (short, max 4 words).\n"
        "2. 'edges': array with 'id', 'source', 'target'.\n"
        "3. 'courses': array of 4 objects with:\n"
        "   - 'title': exact course name\n"
        "   - 'platform': Udemy/Coursera/freeCodeCamp\n"
        "   - 'url': the DIRECT link to the course page (e.g. 'https://www.coursera.org/learn/html-css-javascript-for-web-developers' or 'https://www.udemy.com/course/the-web-developer-bootcamp/'). Must be a real, working URL.\n"
        "   - 'level': Beginner/Intermediate/Advanced\n"
        "   - 'description': 1 short sentence\n"
        "4. 'certificates': array of 3 objects with:\n"
        "   - 'title': exact certification name\n"
        "   - 'provider': Google/AWS/Meta/Microsoft/IBM\n"
        "   - 'url': direct link to the certification page (e.g. 'https://grow.google/certificates/it-support/')\n"
        "   - 'description': 1 short sentence\n"
        "5. 'youtube': array of 4 objects with:\n"
        "   - 'title': exact video title\n"
        "   - 'channel': real channel name (Traversy Media, freeCodeCamp, Fireship, etc.)\n"
        "   - 'videoId': the real 11-character YouTube video ID (e.g. 'hdI2bqOjy3c'). This MUST be a real video ID from YouTube.\n"
        "   - 'description': 1 short sentence\n"
        "CRITICAL: All URLs and videoIds MUST be real and working. Use only well-known, popular resources. "
        "Do not include any text outside the JSON object."
    )

    try:
        completion = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Generate a roadmap for: {skill}"}
            ],
            response_format={"type": "json_object"}
        )
        
        roadmap_data = json.loads(completion.choices[0].message.content)
        
        # Sanitize and fill using shared helper
        roadmap_data = sanitize_roadmap_resources(roadmap_data, skill)
        
        _cache_set(f"roadmap:skill:{skill.lower()}", roadmap_data)
        return roadmap_data
    except Exception as e:
        logger.error(f"Groq API error: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate roadmap")


class RoleLearningPathRequest(BaseModel):
    role: str


@app.post("/api/generate-role-learning-path")
def generate_role_learning_path(data: RoleLearningPathRequest):
    """Generate a comprehensive learning + career roadmap for a target role.

    Returns a strict JSON object with nodes/edges/courses/certificates/youtube similar
    to the skill-level roadmap but scoped to the whole role (career steps, recommended
    months of study, project milestones, first-job checklist).
    """
    role = data.role.strip()
    if not role:
        raise HTTPException(status_code=400, detail="Role name is required")

    cache_key = f"roadmap:role:{role.lower()}"
    cached = _cache_get(cache_key)
    if cached:
        logger.info(f"Serving cached roadmap for role: {role}")
        return cached

    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured")

    client = Groq(api_key=api_key)

    # Use the same strict skill-level roadmap prompt so the returned JSON
    # matches what the frontend `RoadmapViewport` expects (nodes, edges, courses, certificates, youtube).
    system_prompt = (
        "You are a professional curriculum designer. Generate a learning roadmap for the given role. "
        "Return a STRICT JSON object with these keys:\n"
        "1. 'nodes': array of 6-8 objects with 'id' (like 'n1','n2') and 'label' (short, max 4 words).\n"
        "2. 'edges': array with 'id', 'source', 'target'.\n"
        "3. 'courses': array of 4 objects with:\n"
        "   - 'title': exact course name\n"
        "   - 'platform': Udemy/Coursera/freeCodeCamp\n"
        "   - 'url': the DIRECT link to the course page. Must be a real, working URL.\n"
        "   - 'level': Beginner/Intermediate/Advanced\n"
        "   - 'description': 1 short sentence\n"
        "4. 'certificates': array of 3 objects with:\n"
        "   - 'title': exact certification name\n"
        "   - 'provider': Google/AWS/Meta/Microsoft/IBM\n"
        "   - 'url': direct link to the certification page\n"
        "   - 'description': 1 short sentence\n"
        "5. 'youtube': array of 4 objects with:\n"
        "   - 'title': exact video title\n"
        "   - 'channel': real channel name\n"
        "   - 'videoId': the real 11-character YouTube video ID\n"
        "   - 'description': 1 short sentence\n"
        "CRITICAL: All URLs and videoIds MUST be real and working. Use only well-known, popular resources. Do not include any text outside the JSON object."
    )

    try:
        completion = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Generate a roadmap for role: {role}"}
            ],
            response_format={"type": "json_object"}
        )

        roadmap_data = json.loads(completion.choices[0].message.content)

        # Sanitize and fill using shared helper
        roadmap_data = sanitize_roadmap_resources(roadmap_data, role)
        
        _cache_set(f"roadmap:role:{role.lower()}", roadmap_data)
        return roadmap_data
    except Exception as e:
        logger.error(f"Groq role roadmap error: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate role roadmap")


class ResourcesForRolesRequest(BaseModel):
    roles: Optional[List[str]] = None


@app.post("/api/generate-resources-for-roles")
def generate_resources_for_roles(data: ResourcesForRolesRequest):
    """Generate courses/certificates/youtube lists for multiple roles.

    If `roles` is omitted, generate for all cached roles. Returns a mapping
    { role: { courses: [], certificates: [], youtube: [] } } with invalid
    links replaced by safe search fallbacks.
    """
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured")

    client = Groq(api_key=api_key)

    # Determine roles to generate for
    roles = data.roles or list(ROLE_SKILLS_CACHE.keys())
    if not roles:
        # Fallback to a small curated list if nothing available
        roles = [
            "Full Stack Developer",
            "Data Scientist",
            "DevOps Engineer",
            "Mobile Developer",
        ]

    results = {}

    system_prompt = (
        "You are a helpful curator. For the given ROLE, return a STRICT JSON object with these keys:\n"
        "1. 'courses': array of 3 objects with 'title','platform','url','level','description'.\n"
        "2. 'certificates': array of 2 objects with 'title','provider','url','description'.\n"
        "3. 'youtube': array of 3 objects with 'title','channel','videoId','description'.\n"
        "CRITICAL: Provide real, direct URLs and real YouTube video IDs when possible. Do not include any text outside the JSON object."
    )

    for role in roles:
        cache_key = f"resources:{role.strip().lower()}"
        cached = _cache_get(cache_key)
        if cached is not None:
            results[role] = cached
            continue

        try:
            completion = client.chat.completions.create(
                model="llama-3.1-8b-instant",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"Provide resources for role: {role}"},
                ],
                response_format={"type": "json_object"},
                timeout=60,
            )

            payload = json.loads(completion.choices[0].message.content)
            
            # Use shared helper
            sanitized = sanitize_roadmap_resources(payload, role)
            
            _cache_set(cache_key, sanitized)
            results[role] = sanitized
        except Exception as e:
            logger.error("Groq resources error for role %s: %s", role, e)
            results[role] = {"courses": [], "certificates": [], "youtube": []}
            continue

        # store sanitized result in cache
        _cache_set(cache_key, sanitized)
        results[role] = sanitized

    return results


class RoadmapAssistantRequest(BaseModel):
    """Request body for the roadmap assistant chat.

    - question: the latest user question
    - role: optional target role (e.g. Full Stack Developer)
    - skill: optional focus skill (e.g. React)
    - history: optional short list of previous turns as generic dicts
    """

    question: str
    role: Optional[str] = None
    skill: Optional[str] = None
    history: Optional[List[dict]] = None


@app.post("/api/roadmap-assistant")
def roadmap_assistant(data: RoadmapAssistantRequest):
    """Lightweight Groq-powered assistant for roadmap & skill questions.

    Returns a short, on-point natural language answer.
    """

    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not configured")

    client = Groq(api_key=api_key)

    system_prompt = (
        "You are a friendly, concise mentor helping students with tech careers. "
        "Answer in 2-4 short sentences, directly and clearly. "
        "Focus on practical, actionable advice about skills, learning paths, and careers. "
        "Keep the tone encouraging and simple. If the question is off-topic, gently steer it back."
    )

    messages: List[dict] = [{"role": "system", "content": system_prompt}]

    context_bits = []
    if data.role:
        context_bits.append(f"target role: {data.role}")
    if data.skill:
        context_bits.append(f"focus skill: {data.skill}")
    if context_bits:
        messages.append({"role": "system", "content": "Context: " + "; ".join(context_bits)})

    if data.history:
        for turn in data.history[-6:]:  # only keep a short window
            try:
                sender = (turn.get("from") or turn.get("sender") or "user").lower()
                text = (turn.get("text") or "").strip()
            except AttributeError:
                continue
            if not text:
                continue
            role = "assistant" if sender == "assistant" else "user"
            messages.append({"role": role, "content": text})

    messages.append({"role": "user", "content": data.question})

    try:
        completion = client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=messages,
            max_tokens=256,
            temperature=0.4,
        )
        answer = completion.choices[0].message.content.strip()
        return {"answer": answer}
    except Exception as e:
        logger.error(f"Groq assistant error: {e}")
        raise HTTPException(status_code=500, detail="Failed to get assistant answer")


@app.post("/api/analyze-full-profile")
async def analyze_full_profile(
    resume: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    All-in-one endpoint: accepts a PDF resume + usernames via
    multipart/form-data, extracts features, runs the readiness model,
    saves the result to Supabase, and returns a combined response.

    If no resume is uploaded, uses the previously saved resume.
    Requires JWT authentication.
    """
    # ── 1. Extract text from uploaded PDF (or use saved) ──────────────
    resume_text = ""
    resume_filename = ""
    if resume and resume.filename:
        resume_filename = resume.filename
        try:
            contents = await resume.read()
            reader = PdfReader(io.BytesIO(contents))
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    resume_text += page_text + "\n"

            # Persist uploaded file so TPO resume viewer can open it later.
            safe_suffix = Path(resume_filename).name.replace(" ", "_")
            stored_name = f"{current_user.id}_{int(datetime.now(timezone.utc).timestamp())}_{safe_suffix}"
            stored_path = RESUMES_DIR / stored_name
            stored_path.write_bytes(contents)

            current_user.resume_filename = resume_filename
            current_user.resume_url = f"/uploads/resumes/{stored_name}"
            current_user.resume_text = resume_text
        except Exception as exc:
            logger.error("Failed to parse uploaded PDF: %s", exc)
    elif current_user.resume_text:
        # Fall back to saved resume
        resume_text = current_user.resume_text
        resume_filename = current_user.resume_filename or "saved_resume.pdf"

    # ── 2. Extract feature vector ─────────────────────────────────────
    feature_vector = build_complete_feature_vector(
        resume_text=resume_text,
    )

    # ── 3. Run readiness + role prediction using Groq AI ─────────────
    extracted_data = build_structured_candidate_data(
        feature_vector=feature_vector,
        resume_text=resume_text,
    )
    ai_prediction = get_ai_prediction(extracted_data, feature_vector)
    readiness = float(ai_prediction.get("score", 0))

    # ── 4. Categorise readiness ───────────────────────────────────────
    if readiness >= 75:
        category = "Placement Ready"
    elif readiness >= 50:
        category = "Almost Ready"
    else:
        category = "Needs Improvement"

    # ── 5. Role recommendation from AI ───────────────────────────────
    role_list = ai_prediction.get("recommended_roles", [])
    if not role_list and ai_prediction.get("role"):
        role_list = [{"role": ai_prediction["role"], "score": readiness}]

    # ── 6. Save result to Supabase ────────────────────────────────────
    try:
        analysis = AnalysisResult(
            user_id=current_user.id,
            resume_text_preview=resume_text[:200] if resume_text else "",
            github_username="",
            leetcode_username="",
            features=feature_vector,
            readiness_score=readiness,
            readiness_category=category,
            recommended_roles=role_list,
        )
        db.add(analysis)

        # Auto-save resume to user profile
        if resume_text and resume_filename and not current_user.resume_filename:
            current_user.resume_text = resume_text
            current_user.resume_filename = resume_filename

        # Keep the canonical resume score on users table for all dashboards/APIs.
        current_user.resume_score = readiness

        db.commit()
        logger.info("Saved analysis result %s for user %s", analysis.id, current_user.id)
    except Exception as exc:
        logger.error("Failed to save analysis result: %s", exc)
        db.rollback()
        # Don't crash — still return the result to the user

    # ── Debug: Print non-zero features for user visibility ────────────────
    non_zero_features = {k: v for k, v in feature_vector.items() if v > 0}
    print("\n════════════════════════════════════════════════════════════════")
    print(f" ANALYSIS RESULT FOR: {current_user.name}")
    print(f" ----------------------------------------------------------------")
    print(f" READINESS SCORE: {readiness}")
    print(f" DETECTED FEATURES: {len(non_zero_features)}")
    print(f" NON-ZERO VALUES: {non_zero_features}")
    print("════════════════════════════════════════════════════════════════\n")

    return {
        "readiness_score": readiness,
        "readiness_category": category,
        "recommended_roles": role_list,
        "weakness": ai_prediction.get("weakness", []),
        "suggestions": ai_prediction.get("suggestions", []),
        "prediction_source": ai_prediction.get("source", "fallback"),
        "total_features_used": len(non_zero_features),
        "features": non_zero_features,
    }

class PhotoUpdateRequest(BaseModel):
    photo_url: Optional[str] = None
    filePath: Optional[str] = None

@app.post("/api/profile/photo")
@app.post("/api/auth/profile/photo")
def update_profile_photo(
    data: PhotoUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Save the public Supabase photo URL or filePath to the user's profile."""
    photo_url = data.photo_url
    if not photo_url and data.filePath:
        # Construct public URL for profile_photos bucket
        supabase_url = os.getenv("VITE_SUPABASE_URL", "https://swxwubiwmmsezuelrisg.supabase.co")
        photo_url = f"{supabase_url}/storage/v1/object/public/profile_photos/{data.filePath}"
    
    if photo_url:
        current_user.photo_url = photo_url
        db.commit()
    return {"message": "Photo URL updated", "photo_url": photo_url}


# ═══════════════════════════════════════════════════════════════════════════════
# HISTORY ENDPOINT
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/history")
@app.get("/api/history")
def get_history(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return the current user's past analysis results, newest first."""
    results = (
        db.query(AnalysisResult)
        .filter(AnalysisResult.user_id == current_user.id)
        .order_by(AnalysisResult.created_at.desc())
        .limit(20)
        .all()
    )

    return [
        {
            "id": str(r.id),
            "github_username": r.github_username,
            "leetcode_username": r.leetcode_username,
            "readiness_score": r.readiness_score,
            "readiness_category": r.readiness_category,
            "recommended_roles": r.recommended_roles,
            "education_score": r.education_score,
            "skills_score": r.skills_score,
            "contact_score": r.contact_score,
            "internship_score": r.internship_score,
            "experience_score": r.experience_score,
            "project_score": r.project_score,
            "ai_suggestions": r.ai_suggestions,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in results
    ]


@app.post("/analyze-student")
@app.post("/api/analyze-student")
def analyze_student(data: StudentFeatures):
    normalized_features = dict(data.features or {})
    for col in feature_columns:
        if col not in normalized_features:
            normalized_features[col] = 0

    extracted_data = build_structured_candidate_data(
        feature_vector=normalized_features,
        resume_text="",
    )
    ai_prediction = get_ai_prediction(extracted_data, normalized_features)
    readiness = float(ai_prediction.get("score", 0))

    top_roles = ai_prediction.get("recommended_roles", [])
    if not top_roles and ai_prediction.get("role"):
        top_roles = [{"role": ai_prediction["role"], "score": readiness}]

    return {
        "readiness_score": round(readiness, 2),
        "recommended_roles": top_roles,
        "weakness": ai_prediction.get("weakness", []),
        "suggestions": ai_prediction.get("suggestions", []),
        "prediction_source": ai_prediction.get("source", "fallback"),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# QUIZ ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

class QuizGenerateRequest(BaseModel):
    role: str
    difficulty: str
    resultId: Optional[str] = None

class QuizSubmitRequest(BaseModel):
    role: str
    difficulty: str
    score: int
    totalQuestions: int
    questions: list # Array of stored question objects
    answers: list # List of objects
    resultId: Optional[str] = None

@app.post("/quiz/generate")
@app.post("/api/quiz/generate")
def generate_quiz_endpoint(
    data: QuizGenerateRequest, 
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        if data.resultId:
            # Retest scenario: fetch the exact previous questions
            result = db.query(QuizResult).filter(
                QuizResult.id == data.resultId,
                QuizResult.user_id == current_user.id
            ).first()
            if result and result.questions:
                return {"questions": result.questions}
                
        # New quiz scenario: generate fresh questions
        questions = generate_quiz_questions(data.role, data.difficulty)
        return {"questions": questions}
    except ValueError as e:
        logger.error("Quiz generation failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.error("Quiz generation failed: %s", e)
        raise HTTPException(status_code=500, detail="Failed to generate quiz")

@app.post("/quiz/submit")
@app.post("/api/quiz/submit")
def submit_quiz_endpoint(
    data: QuizSubmitRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # Check if result_id is provided for retest (updates existing result)
        if data.resultId:
            result = db.query(QuizResult).filter(
                QuizResult.id == data.resultId, 
                QuizResult.user_id == current_user.id
            ).first()
            
            if result:
                result.role = data.role
                result.difficulty = data.difficulty
                result.score = data.score
                result.total_questions = data.totalQuestions
                # We do NOT store questions and answers in the DB anymore
                result.questions = [] 
                result.answers = []
                db.commit()
                db.refresh(result)
                return {"message": "Quiz result updated successfully", "resultId": str(result.id)}
        
        # New submission
        result = QuizResult(
            user_id=current_user.id,
            role=data.role,
            difficulty=data.difficulty,
            score=data.score,
            total_questions=data.totalQuestions,
            questions=[], # Do not store detailed questions
            answers=[]    # Do not store detailed answers
        )
        db.add(result)
        db.commit()
        db.refresh(result)
        return {"message": "Quiz submitted successfully", "resultId": str(result.id)}
    except Exception as e:
        logger.error("Quiz submit failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/quiz/results")
@app.get("/api/quiz/results")
def get_quiz_results(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    results = db.query(QuizResult).filter(QuizResult.user_id == current_user.id).order_by(QuizResult.created_at.desc()).all()
    return {"results": results}

@app.get("/quiz/roles")
@app.get("/api/quiz/roles")
def get_quiz_roles(current_user: User = Depends(get_current_user)):
     roles = [
        'Backend Developer', 'Frontend Developer', 'Full Stack Developer',
        'ML Engineer', 'Data Scientist', 'Data Engineer',
        'Java Developer', 'Python Developer', 'DevOps Engineer',
        'Cloud Engineer', 'Mobile Developer', 'iOS Developer',
        'Android Developer', 'QA / Test Engineer', 'Cybersecurity Analyst',
        'AI Research Engineer', 'Game Developer', 'Blockchain Developer',
        'Database Administrator', 'Systems Engineer', 'UI/UX Designer',
    ]
     return {"roles": roles}


# ══════════════════════════════════════════════════════════════════════════════
# TPO  —  Job-posting & applicant management
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/api/tpo/jobs")
async def create_job(
    title: str = Form(...),
    company: str = Form(...),
    description: str = Form(""),
    eligibility: str = Form(""),
    job_role: str = Form(""),
    min_cgpa: Optional[float] = Form(None),
    min_resume_score: Optional[float] = Form(None),
    required_certifications: str = Form(""),
    preferred_skills: str = Form(""),
    package_lpa: Optional[float] = Form(None),
    deadline: str = Form(""),
    company_logo: UploadFile = File(None),
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """TPO creates a new job posting (multipart/form-data with optional logo)."""
    logo_data = ""
    if company_logo and company_logo.filename:
        allowed = (".png", ".jpg", ".jpeg")
        ext = Path(company_logo.filename).suffix.lower()
        if ext not in allowed:
            raise HTTPException(status_code=400, detail=f"Logo must be PNG, JPG, or JPEG. Got: {ext}")
        raw = await company_logo.read()
        mime = "image/png" if ext == ".png" else "image/jpeg"
        logo_data = f"data:{mime};base64,{base64.b64encode(raw).decode()}"

    normalized_min_resume_score = None
    if min_resume_score is not None:
        normalized_min_resume_score = max(0.0, min(100.0, float(min_resume_score)))

    job = Job(
        posted_by=tpo.id,
        title=title,
        company=company,
        description=description,
        eligibility=eligibility,
        job_role=job_role,
        min_cgpa=min_cgpa,
        min_resume_score=normalized_min_resume_score,
        required_certifications=required_certifications,
        preferred_skills=preferred_skills,
        package_lpa=package_lpa,
        deadline=deadline,
        company_logo=logo_data,
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return {
        "id": str(job.id),
        "title": job.title,
        "company": job.company,
        "description": job.description,
        "eligibility": job.eligibility,
        "job_role": job.job_role,
        "min_cgpa": job.min_cgpa,
        "min_resume_score": job.min_resume_score,
        "required_certifications": job.required_certifications,
        "preferred_skills": job.preferred_skills,
        "package_lpa": job.package_lpa,
        "deadline": job.deadline,
        "company_logo": job.company_logo or "",
        "created_at": str(job.created_at),
    }


@app.get("/api/tpo/jobs")
def list_tpo_jobs(
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """List all jobs posted by this TPO."""
    jobs = (
        db.query(Job)
        .filter(Job.posted_by == tpo.id)
        .order_by(Job.created_at.desc())
        .all()
    )
    return {
        "jobs": [
            {
                "id": str(j.id),
                "title": j.title,
                "company": j.company,
                "description": j.description,
                "eligibility": j.eligibility,
                "job_role": j.job_role or "",
                "min_cgpa": j.min_cgpa,
                "min_resume_score": j.min_resume_score,
                "required_certifications": j.required_certifications or "",
                "preferred_skills": j.preferred_skills or "",
                "package_lpa": j.package_lpa,
                "deadline": j.deadline,
                "company_logo": j.company_logo or "",
                "created_at": str(j.created_at),
            }
            for j in jobs
        ]
    }


@app.delete("/api/tpo/jobs/{job_id}")
def delete_job(
    job_id: str,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """TPO deletes one of their own job postings."""
    job = db.query(Job).filter(Job.id == job_id, Job.posted_by == tpo.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    db.delete(job)
    db.commit()
    return {"detail": "Job deleted"}


@app.post("/api/results/upload")
async def upload_job_results(
    job_id: str = Form(...),
    round_name: str = Form(...),
    status_value: str = Form(..., alias="status"),
    remarks: str = Form(""),
    result_file: UploadFile = File(None),
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Upload round-wise results to all shortlisted + interested students for a TPO-owned job."""
    normalized_round_name = (round_name or "").strip()
    if not normalized_round_name:
        raise HTTPException(status_code=400, detail="round_name is required")

    normalized_status = (status_value or "").strip().title()
    if normalized_status not in RESULT_STATUSES:
        raise HTTPException(status_code=400, detail="status must be one of Selected, Rejected, or Qualified")

    job = db.query(Job).filter(Job.id == job_id, Job.posted_by == tpo.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    shortlisted_ids = {
        str(row.student_id)
        for row in db.query(ShortlistedJob.student_id).filter(ShortlistedJob.job_id == job_id).all()
    }
    interested_ids = {
        str(row.student_id)
        for row in db.query(InterestedJob.student_id).filter(InterestedJob.job_id == job_id).all()
    }
    valid_ids = sorted(shortlisted_ids.union(interested_ids))

    if not valid_ids:
        raise HTTPException(
            status_code=400,
            detail="No shortlisted or interested students found for this job",
        )

    normalized_remarks = (remarks or "").strip()
    student_rows = [sid for sid in valid_ids]

    file_url = ""
    if result_file and result_file.filename:
        ext = Path(result_file.filename).suffix.lower()
        if ext not in {".pdf", ".xlsx", ".xls"}:
            raise HTTPException(status_code=400, detail="Result file must be PDF or Excel (.xlsx/.xls)")

        safe_round = re.sub(r"[^a-zA-Z0-9_-]+", "_", normalized_round_name).strip("_") or "round"
        storage_name = f"{job_id}_{int(time.time())}_{safe_round}{ext}"
        destination = RESULTS_DIR / storage_name
        content = await result_file.read()
        destination.write_bytes(content)
        file_url = f"/uploads/results/{storage_name}"

    result = JobResult(
        job_id=job_id,
        round_name=normalized_round_name,
        result_status=normalized_status,
        remarks=normalized_remarks,
        students=student_rows,
        file_url=file_url,
    )
    db.add(result)

    for sid in valid_ids:
        db.add(
            Notification(
                student_id=sid,
                message=f"Result update for {job.company} - {job.title} ({normalized_round_name}): {normalized_status}.",
            )
        )

    db.commit()
    db.refresh(result)

    return {
        "message": "Results uploaded successfully",
        "result": {
            "id": str(result.id),
            "job_id": str(result.job_id),
            "round_name": result.round_name,
            "status": result.result_status,
            "remarks": result.remarks,
            "students": result.students or [],
            "file_url": result.file_url or "",
            "created_at": str(result.created_at),
        },
    }


@app.get("/api/results/{job_id}")
def get_job_results(
    job_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get uploaded results for a job with role-based filtering."""
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    all_results = (
        db.query(JobResult)
        .filter(JobResult.job_id == job_id)
        .order_by(JobResult.created_at.desc())
        .all()
    )

    job_payload = {
        "id": str(job.id),
        "title": job.title,
        "company": job.company,
    }

    if current_user.role == "tpo":
        if str(job.posted_by) != str(current_user.id):
            raise HTTPException(status_code=403, detail="You can only view results for your own jobs")

        return {
            "job": job_payload,
            "results": [
                {
                    "id": str(item.id),
                    "round_name": item.round_name,
                    "status": item.result_status or "",
                    "remarks": item.remarks or "",
                    "students": item.students or [],
                    "file_url": item.file_url or "",
                    "created_at": str(item.created_at),
                }
                for item in all_results
            ],
        }

    student_id = str(current_user.id)
    is_interested = db.query(InterestedJob.id).filter(
        InterestedJob.job_id == job_id,
        InterestedJob.student_id == student_id,
    ).first()
    is_shortlisted = db.query(ShortlistedJob.id).filter(
        ShortlistedJob.job_id == job_id,
        ShortlistedJob.student_id == student_id,
    ).first()

    if not (is_interested or is_shortlisted):
        return {"job": job_payload, "results": []}

    filtered_results = []
    for item in all_results:
        entries = item.students or []
        row_status = item.result_status or ""
        row_remarks = item.remarks or ""

        is_targeted = False
        if entries and isinstance(entries[0], dict):
            student_row = next(
                (
                    row
                    for row in entries
                    if str(row.get("studentId") or row.get("student_id") or "").strip() == student_id
                ),
                None,
            )
            if student_row:
                is_targeted = True
                row_status = student_row.get("status", row_status)
                row_remarks = student_row.get("remarks", row_remarks)
        else:
            targeted_ids = {str(sid).strip() for sid in entries}
            is_targeted = student_id in targeted_ids

        if not is_targeted:
            continue

        filtered_results.append(
            {
                "id": str(item.id),
                "round_name": item.round_name,
                "status": row_status,
                "remarks": row_remarks,
                "file_url": item.file_url or "",
                "created_at": str(item.created_at),
            }
        )

    return {"job": job_payload, "results": filtered_results}


@app.get("/api/tpo/jobs/{job_id}/shortlisted")
def get_shortlisted_students(
    job_id: str,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """
    Return auto-shortlisted and manually-shortlisted students for this job.
    Both sections are sourced from shortlisted_jobs using the `source` column.
    """
    job = db.query(Job).filter(Job.id == job_id, Job.posted_by == tpo.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    req_certs = [
        c.strip().lower()
        for c in (job.required_certifications or "").split(",")
        if c.strip()
    ]

    # Parse preferred skills (lowercased, trimmed)
    pref_skills = [
        s.strip().lower()
        for s in (job.preferred_skills or "").split(",")
        if s.strip()
    ]

    existing_shortlists = (
        db.query(ShortlistedJob)
        .filter(ShortlistedJob.job_id == job_id)
        .all()
    )
    existing_by_student = {str(row.student_id): row for row in existing_shortlists}

    students = db.query(User).filter(User.role == "student").all()

    def normalize_resume_url_for_response(raw_url: Optional[str]) -> str:
        """Return a stable resume URL that works for both legacy and current stored values."""
        url = (raw_url or "").strip()
        if not url:
            return ""

        # Keep public external links untouched.
        if url.startswith("http://") or url.startswith("https://"):
            parsed = urlparse(url)
            if parsed.path.startswith("/uploads/") and parsed.hostname in {"localhost", "127.0.0.1"}:
                return parsed.path
            return url

        normalized = url.replace("\\", "/")
        lower = normalized.lower()

        # Legacy values may contain absolute filesystem paths. Extract app-relative uploads path.
        uploads_marker = "/uploads/"
        idx = lower.find(uploads_marker)
        if idx != -1:
            return normalized[idx:]

        # Older records could store /resumes/... without the /uploads prefix.
        resumes_marker = "/resumes/"
        ridx = lower.find(resumes_marker)
        if ridx != -1:
            return f"/uploads{normalized[ridx:]}"

        # Last fallback: if we can identify a filename, point to canonical resumes location.
        filename = Path(normalized).name
        if filename:
            candidate = RESUMES_DIR / filename
            if candidate.exists():
                return f"/uploads/resumes/{filename}"

        return normalized if normalized.startswith("/") else f"/{normalized}"

    def build_shortlist_item(s: User, source: str) -> dict:
        resume_sc = s.resume_score or 0

        # Compute informational match details for display.
        student_text = " ".join([
            (s.resume_text or "").lower(),
            (s.certifications or "").lower(),
            (s.preferred_job_roles or "").lower(),
        ])
        if pref_skills:
            matched_skills = [sk for sk in pref_skills if sk in student_text]
            skill_match = len(matched_skills) / len(pref_skills)
        else:
            matched_skills = []
            skill_match = 1.0

        student_certs = [
            c.strip().lower()
            for c in (s.certifications or "").split(",")
            if c.strip()
        ]
        matched_certs = [rc for rc in req_certs if rc in student_certs] if req_certs else []

        cgpa_score = min((s.cgpa or 0) / 10.0, 1.0)
        match_score = round(((skill_match + cgpa_score + resume_sc / 100) / 3) * 100, 2)

        normalized_resume_url = normalize_resume_url_for_response(s.resume_url)
        student_payload = {
            "id": str(s.id),
            "name": s.name,
            "email": s.email,
            "mobile_number": s.mobile_number or "",
            "cgpa": s.cgpa,
            "certifications": s.certifications or "",
            "preferred_job_roles": s.preferred_job_roles or "",
            "resume_text": s.resume_text or "",
            "resume_score": resume_sc,
            "resume_url": normalized_resume_url,
        }

        return {
            "student": student_payload,
            "id": student_payload["id"],
            "name": student_payload["name"],
            "email": student_payload["email"],
            "cgpa": student_payload["cgpa"],
            "resume_score": student_payload["resume_score"],
            "resume_url": student_payload["resume_url"],
            "match_score": match_score,
            "matched_skills": matched_skills,
            "matched_certifications": matched_certs,
            "source": source,
        }

    required_min_resume = max(0.0, min(100.0, float(job.min_resume_score or 0)))
    eligible_auto_ids = set()
    for s in students:
        sid = str(s.id)
        # Manual rows should remain manual and never be overwritten to auto.
        if sid in existing_by_student and (existing_by_student[sid].source or "manual") == "manual":
            continue

        if job.min_cgpa is not None and job.min_cgpa > 0:
            if s.cgpa is None or s.cgpa < job.min_cgpa:
                continue

        resume_sc = s.resume_score or 0
        if resume_sc < required_min_resume:
            continue

        eligible_auto_ids.add(sid)

    # Sync auto rows in shortlisted_jobs so source-based queries are accurate.
    existing_auto_ids = {
        sid
        for sid, row in existing_by_student.items()
        if (row.source or "manual") == "auto"
    }

    auto_ids_to_add = eligible_auto_ids - set(existing_by_student.keys())
    for sid in auto_ids_to_add:
        db.add(ShortlistedJob(job_id=job_id, student_id=sid, source="auto"))

    auto_ids_to_remove = existing_auto_ids - eligible_auto_ids
    if auto_ids_to_remove:
        db.query(ShortlistedJob).filter(
            ShortlistedJob.job_id == job_id,
            ShortlistedJob.source == "auto",
            ShortlistedJob.student_id.in_(list(auto_ids_to_remove)),
        ).delete(synchronize_session=False)

    if auto_ids_to_add or auto_ids_to_remove:
        db.commit()

    auto_shortlisted_pairs = (
        db.query(ShortlistedJob, User)
        .join(User, User.id == ShortlistedJob.student_id)
        .filter(ShortlistedJob.job_id == job_id, ShortlistedJob.source == "auto")
        .order_by(ShortlistedJob.created_at.desc())
        .all()
    )
    manual_shortlisted_pairs = (
        db.query(ShortlistedJob, User)
        .join(User, User.id == ShortlistedJob.student_id)
        .filter(
            ShortlistedJob.job_id == job_id,
            or_(
                ShortlistedJob.source == "manual",
                ShortlistedJob.source.is_(None),
                ShortlistedJob.source == "",
            ),
        )
        .order_by(ShortlistedJob.created_at.desc())
        .all()
    )

    auto_shortlisted = [
        build_shortlist_item(user, source="auto")
        for _, user in auto_shortlisted_pairs
    ]
    auto_shortlisted_ids = {item["id"] for item in auto_shortlisted}
    manual_shortlisted = [
        build_shortlist_item(user, source="manual")
        for _, user in manual_shortlisted_pairs
        if str(user.id) not in auto_shortlisted_ids
    ]

    auto_shortlisted.sort(key=lambda x: x["match_score"], reverse=True)
    shortlisted = auto_shortlisted + manual_shortlisted

    return {
        "job": {
            "id": str(job.id),
            "title": job.title,
            "company": job.company,
            "min_cgpa": job.min_cgpa,
            "min_resume_score": job.min_resume_score,
            "required_certifications": job.required_certifications or "",
            "preferred_skills": job.preferred_skills or "",
            "job_role": job.job_role or "",
        },
        "auto_shortlisted_students": auto_shortlisted,
        "manual_shortlisted_students": manual_shortlisted,
        "shortlisted_students": shortlisted,
        "total": len(shortlisted),
    }


@app.post("/api/jobs/shortlist-student")
def shortlist_student(
    payload: ShortlistStudentRequest,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Shortlist one student for a job and remove them from interested list."""
    job = db.query(Job).filter(Job.id == payload.job_id, Job.posted_by == tpo.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    student = db.query(User).filter(User.id == payload.student_id, User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    existing_shortlist = db.query(ShortlistedJob).filter(
        ShortlistedJob.job_id == payload.job_id,
        ShortlistedJob.student_id == payload.student_id,
    ).first()
    if existing_shortlist:
        return {
            "message": "Student already shortlisted",
            "job_id": payload.job_id,
            "student_id": payload.student_id,
            "already_shortlisted": True,
        }

    db.add(ShortlistedJob(job_id=payload.job_id, student_id=payload.student_id, source="manual"))

    existing_interest = db.query(InterestedJob).filter(
        InterestedJob.job_id == payload.job_id,
        InterestedJob.student_id == payload.student_id,
    ).first()
    if existing_interest:
        db.delete(existing_interest)

    db.commit()
    return {"message": "Student shortlisted", "job_id": payload.job_id, "student_id": payload.student_id}


@app.post("/api/jobs/shortlist-all")
def shortlist_all_students(
    payload: ShortlistAllRequest,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Shortlist selected students for a job and remove them from interested list."""
    job = db.query(Job).filter(Job.id == payload.job_id, Job.posted_by == tpo.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    student_ids = [sid for sid in payload.student_ids if sid]
    if not student_ids:
        raise HTTPException(status_code=400, detail="student_ids is required")

    valid_students = db.query(User.id).filter(User.id.in_(student_ids), User.role == "student").all()
    valid_ids = {str(row.id) for row in valid_students}
    if not valid_ids:
        raise HTTPException(status_code=400, detail="No valid students provided")

    existing_shortlisted_rows = {
        str(row.student_id): row
        for row in db.query(ShortlistedJob).filter(
            ShortlistedJob.job_id == payload.job_id,
            ShortlistedJob.student_id.in_(list(valid_ids)),
        ).all()
    }

    inserted_ids: List[str] = []
    skipped_ids: List[str] = []
    for sid in valid_ids:
        existing_row = existing_shortlisted_rows.get(sid)
        if not existing_row:
            db.add(ShortlistedJob(job_id=payload.job_id, student_id=sid, source="manual"))
            inserted_ids.append(sid)
        else:
            skipped_ids.append(sid)

    db.query(InterestedJob).filter(
        InterestedJob.job_id == payload.job_id,
        InterestedJob.student_id.in_(inserted_ids),
    ).delete(synchronize_session=False)

    db.commit()
    if inserted_ids:
        message = "Students shortlisted"
    else:
        message = "All selected students are already shortlisted"

    return {
        "message": message,
        "job_id": payload.job_id,
        "student_ids": inserted_ids,
        "already_shortlisted_ids": skipped_ids,
    }


def _build_resumes_zip(students: List[dict], archive_name: str) -> StreamingResponse:
    """Create an in-memory zip for student resumes and return as download response."""
    if not students:
        raise HTTPException(status_code=404, detail="No students found")

    zip_buffer = io.BytesIO()
    added_files = 0
    with ZipFile(zip_buffer, mode="w", compression=ZIP_DEFLATED) as zip_file:
        for s in students:
            resume_url = (s.get("resume_url") or "").strip()
            if not resume_url:
                continue
            relative_resume_path = resume_url.lstrip("/")
            resume_path = BASE_DIR / relative_resume_path
            if not resume_path.exists() or not resume_path.is_file():
                continue

            student_name = (s.get("name") or "student").strip()
            safe_student_name = re.sub(r"[^a-zA-Z0-9_-]+", "_", student_name).strip("_") or "student"
            extension = resume_path.suffix or ".pdf"
            arcname = f"{safe_student_name}_{s.get('id', '')}{extension}"
            zip_file.write(resume_path, arcname=arcname)
            added_files = added_files + 1

    if added_files == 0:
        raise HTTPException(status_code=404, detail="No resume files found for the selected students")

    zip_buffer.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{archive_name}"'}
    return StreamingResponse(zip_buffer, media_type="application/zip", headers=headers)


@app.get("/api/resumes/download-all")
def download_all_shortlisted_resumes(
    job_id: str,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Download all resumes from shortlisted students of the selected job as a ZIP."""
    shortlisted_data = get_shortlisted_students(job_id=job_id, tpo=tpo, db=db)
    shortlisted_students = [item.get("student", {}) for item in shortlisted_data.get("shortlisted_students", [])]
    return _build_resumes_zip(shortlisted_students, archive_name="shortlisted_all_resumes.zip")


@app.post("/api/resumes/download-selected")
def download_selected_shortlisted_resumes(
    payload: DownloadSelectedResumesRequest,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Download selected students' resumes as ZIP, optionally restricted to shortlisted students for a job."""
    selected_ids = [str(sid) for sid in (payload.student_ids or []) if str(sid).strip()]
    if not selected_ids:
        raise HTTPException(status_code=400, detail="student_ids is required")

    selected_set = set(selected_ids)
    students: List[dict] = []

    if payload.job_id:
        shortlisted_data = get_shortlisted_students(job_id=payload.job_id, tpo=tpo, db=db)
        shortlisted_students = [item.get("student", {}) for item in shortlisted_data.get("shortlisted_students", [])]
        students = [s for s in shortlisted_students if str(s.get("id")) in selected_set]
    else:
        db_students = (
            db.query(User)
            .filter(User.id.in_(selected_ids), User.role == "student")
            .all()
        )
        students = [
            {
                "id": str(s.id),
                "name": s.name,
                "resume_url": s.resume_url or "",
            }
            for s in db_students
        ]

    return _build_resumes_zip(students, archive_name="shortlisted_selected_resumes.zip")


@app.get("/api/tpo/jobs/{job_id}/shortlisted/export")
def export_shortlisted_students(
    job_id: str,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Export shortlisted students as Excel with job details and student details."""
    wb = Workbook()

    job = db.query(Job).filter(Job.id == job_id, Job.posted_by == tpo.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    shortlisted_data = get_shortlisted_students(job_id=job_id, tpo=tpo, db=db)
    shortlisted_students = shortlisted_data.get("shortlisted_students", [])
    lpa_value = job.package_lpa
    if lpa_value is None:
        lpa_value = db.execute(
            text("SELECT salary FROM jobs WHERE id = :job_id"),
            {"job_id": job_id},
        ).scalar()

    wb = Workbook()
    ws = wb.active
    ws.title = "Shortlisted Students"

    headers = [
        "Job Title",
        "Company",
        "Student Name",
        "Email",
        "Phone",
        "CGPA",
        "Resume Score",
        "Match Percentage",
        "Certifications",
        "Preferred Role",
    ]
    ws.append(headers)

    for item in shortlisted_students:
        student = item.get("student", {})
        ws.append([
            job.title,
            job.company,
            student.get("name", ""),
            student.get("email", ""),
            student.get("mobile_number", ""),
            student.get("cgpa", ""),
            student.get("resume_score", ""),
            item.get("match_score", ""),
            student.get("certifications", ""),
            student.get("preferred_job_roles", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"shortlisted_students_{job.company}_{job.title}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/students/{student_id}/resume")
def get_student_resume(
    student_id: str,
    download: bool = False,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Return a shortlisted student's uploaded resume PDF for TPO viewing/downloading."""
    _ = tpo  # Dependency enforces TPO access.
    student = db.query(User).filter(User.id == student_id, User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    resume_url = (student.resume_url or "").strip()
    if not resume_url:
        raise HTTPException(status_code=404, detail="Resume not uploaded")

    relative_resume_path = resume_url.lstrip("/")
    resume_path = BASE_DIR / relative_resume_path
    if not resume_path.exists() or not resume_path.is_file():
        raise HTTPException(status_code=404, detail="Resume not uploaded")

    download_name = (student.resume_filename or f"{student.name}_resume.pdf").strip() or "resume.pdf"
    disposition = "attachment" if download else "inline"
    headers = {"Content-Disposition": f'{disposition}; filename="{download_name}"'}
    return FileResponse(path=resume_path, media_type="application/pdf", headers=headers)


# ══════════════════════════════════════════════════════════════════════════════
# Students  —  Browse & apply to jobs
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/jobs")
def list_all_jobs(
    current_user: User = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Students browse all available job postings (view-only, no apply)."""
    jobs = db.query(Job).order_by(Job.created_at.desc()).all()

    return {
        "jobs": [
            {
                "id": str(j.id),
                "title": j.title,
                "company": j.company,
                "description": j.description,
                "eligibility": j.eligibility,
                "job_role": j.job_role or "",
                "min_cgpa": j.min_cgpa,
                "min_resume_score": j.min_resume_score,
                "required_certifications": j.required_certifications or "",
                "preferred_skills": j.preferred_skills or "",
                "package_lpa": j.package_lpa,
                "deadline": j.deadline,
                "company_logo": j.company_logo or "",
                "created_at": str(j.created_at),
            }
            for j in jobs
        ]
    }


# ══════════════════════════════════════════════════════════════════════════════
# INTERESTED JOBS
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/api/jobs/{job_id}/interest")
def mark_interest(
    job_id: str,
    current_user: User = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Student marks interest in a job."""
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    existing = db.query(InterestedJob).filter(
        InterestedJob.student_id == current_user.id,
        InterestedJob.job_id == job_id,
    ).first()
    if existing:
        return {"message": "Already interested", "interested": True}

    interest = InterestedJob(student_id=current_user.id, job_id=job_id)
    db.add(interest)
    db.commit()
    return {"message": "Interest marked", "interested": True}


@app.delete("/api/jobs/{job_id}/interest")
def remove_interest(
    job_id: str,
    current_user: User = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Student removes interest in a job."""
    existing = db.query(InterestedJob).filter(
        InterestedJob.student_id == current_user.id,
        InterestedJob.job_id == job_id,
    ).first()
    if existing:
        db.delete(existing)
        db.commit()
    return {"message": "Interest removed", "interested": False}


@app.get("/api/jobs/interests")
def get_my_interests(
    current_user: User = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Get list of job IDs the current student is interested in."""
    interests = db.query(InterestedJob.job_id).filter(
        InterestedJob.student_id == current_user.id,
    ).all()
    return {"job_ids": [str(i.job_id) for i in interests]}


@app.get("/api/jobs/my-applications")
def get_my_applied_jobs(
    current_user: User = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Return jobs where the student is interested or shortlisted."""
    interested_ids = {
        str(row.job_id)
        for row in db.query(InterestedJob.job_id).filter(InterestedJob.student_id == current_user.id).all()
    }
    shortlisted_ids = {
        str(row.job_id)
        for row in db.query(ShortlistedJob.job_id).filter(ShortlistedJob.student_id == current_user.id).all()
    }

    all_job_ids = interested_ids.union(shortlisted_ids)
    if not all_job_ids:
        return {"jobs": []}

    jobs = (
        db.query(Job)
        .filter(Job.id.in_(list(all_job_ids)))
        .order_by(Job.created_at.desc())
        .all()
    )

    return {
        "jobs": [
            {
                "id": str(j.id),
                "title": j.title,
                "company": j.company,
                "description": j.description,
                "job_role": j.job_role or "",
                "package_lpa": j.package_lpa,
                "deadline": j.deadline,
                "company_logo": j.company_logo or "",
                "is_interested": str(j.id) in interested_ids,
                "is_shortlisted": str(j.id) in shortlisted_ids,
            }
            for j in jobs
        ]
    }


@app.get("/api/tpo/jobs/{job_id}/interested-students")
def get_interested_students(
    job_id: str,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """TPO views students who clicked 'Interested' for a job."""
    job = db.query(Job).filter(Job.id == job_id, Job.posted_by == tpo.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    interests = db.query(InterestedJob).filter(InterestedJob.job_id == job_id).all()
    students = []
    for i in interests:
        s = db.query(User).filter(User.id == i.student_id).first()
        if s:
            students.append({
                "id": str(s.id),
                "name": s.name,
                "moodle_id": s.moodle_id or "",
                "division": s.division or "",
                "email": s.email,
                "resume_score": s.resume_score or 0,
                "cgpa": s.cgpa or 0,
                "interested_at": str(i.created_at),
            })

    return {
        "job": {"id": str(job.id), "title": job.title, "company": job.company},
        "students": students,
        "total": len(students),
    }


@app.get("/api/tpo/jobs/{job_id}/interested-students/export")
def export_interested_students(
    job_id: str,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Export interested students as Excel (.xlsx)."""
    from openpyxl import Workbook

    job = db.query(Job).filter(Job.id == job_id, Job.posted_by == tpo.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    interests = db.query(InterestedJob).filter(InterestedJob.job_id == job_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Interested Students"
    ws.append(["Student Name", "Roll Number", "Department", "Email", "Resume Score", "CGPA"])

    for i in interests:
        s = db.query(User).filter(User.id == i.student_id).first()
        if s:
            ws.append([
                s.name,
                s.moodle_id or "",
                s.division or "",
                s.email,
                s.resume_score or 0,
                s.cgpa or 0,
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"interested_students_{job.company}_{job.title}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# NOTIFICATIONS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/notifications")
def get_notifications(
    current_user: User = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Get student's notifications, newest first."""
    notifs = (
        db.query(Notification)
        .filter(Notification.student_id == current_user.id)
        .order_by(Notification.created_at.desc())
        .limit(50)
        .all()
    )
    return {
        "notifications": [
            {
                "id": str(n.id),
                "message": n.message,
                "status": n.status,
                "created_at": str(n.created_at),
            }
            for n in notifs
        ]
    }


@app.get("/api/notifications/unread-count")
def get_unread_count(
    current_user: User = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Get count of unread notifications."""
    count = db.query(Notification).filter(
        Notification.student_id == current_user.id,
        Notification.status == "unread",
    ).count()
    return {"count": count}


@app.put("/api/notifications/{notification_id}/read")
def mark_notification_read(
    notification_id: str,
    current_user: User = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Mark a notification as read."""
    notif = db.query(Notification).filter(
        Notification.id == notification_id,
        Notification.student_id == current_user.id,
    ).first()
    if not notif:
        raise HTTPException(status_code=404, detail="Notification not found")
    notif.status = "read"
    db.commit()
    return {"message": "Notification marked as read"}


@app.post("/api/tpo/jobs/{job_id}/notify-shortlisted")
def notify_shortlisted(
    job_id: str,
    payload: Optional[NotifyShortlistedRequest] = None,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Send notifications to shortlisted students (all or selected IDs)."""
    job = db.query(Job).filter(Job.id == job_id, Job.posted_by == tpo.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if payload and payload.job_id and payload.job_id != job_id:
        raise HTTPException(status_code=400, detail="Job ID mismatch in request payload")

    shortlisted_data = get_shortlisted_students(job_id=job_id, tpo=tpo, db=db)
    shortlisted_students = shortlisted_data.get("shortlisted_students", [])
    allowed_student_ids = {item["student"]["id"] for item in shortlisted_students}

    selected_ids = set(payload.student_ids) if payload and payload.student_ids else None
    if selected_ids is not None:
        selected_ids = {sid for sid in selected_ids if sid in allowed_student_ids}
        if len(selected_ids) == 0:
            raise HTTPException(status_code=400, detail="No valid selected students found for this job")

    notified_count = 0

    for item in shortlisted_students:
        student = item.get("student", {})
        student_id = student.get("id")
        if not student_id:
            continue

        if selected_ids is not None and student_id not in selected_ids:
            continue

        db_student = db.query(User).filter(User.id == student_id).first()
        if not db_student:
            continue

        # Create notification
        message = f"Congratulations! You have been shortlisted for the {job.job_role or job.title} role at {job.company}."
        notif = Notification(
            student_id=db_student.id,
            message=message,
        )
        db.add(notif)
        notified_count = notified_count + 1

    db.commit()
    return {"message": f"Notifications sent to {notified_count} students.", "count": notified_count}


@app.get("/api/jobs/{job_id}/shortlisted/export")
def export_shortlisted_students_json(
    job_id: str,
    tpo: User = Depends(get_current_tpo),
    db: Session = Depends(get_db),
):
    """Return shortlisted students data in JSON format for export workflows."""
    shortlisted_data = get_shortlisted_students(job_id=job_id, tpo=tpo, db=db)
    rows = []
    for item in shortlisted_data.get("shortlisted_students", []):
        student = item.get("student", {})
        rows.append({
            "student_name": student.get("name", ""),
            "email": student.get("email", ""),
            "phone": student.get("mobile_number", ""),
            "cgpa": student.get("cgpa"),
            "resume_score": student.get("resume_score"),
            "match_percentage": item.get("match_score"),
            "certifications": student.get("certifications", ""),
            "preferred_role": student.get("preferred_job_roles", ""),
        })

    return {
        "job": shortlisted_data.get("job", {}),
        "total": len(rows),
        "students": rows,
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=False)


